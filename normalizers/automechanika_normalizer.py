"""
normalizers/automechanika_normalizer.py

Maps the Ifor Williams "Automechanika Frankfurt exhibitor" export onto
the standard supplier candidate shape, so it can go through
`pipeline.static_list_import.import_static_supplier_list` exactly like
a live scraper's results do.

The exact source file
------------------------
Three sheets share an identical column layout — `name`, `tier`,
`final_score`, `hall_stand`, `website`, `address`,
`matched_product_groups`, `matched_text_keywords`, `description`,
`url` — and differ only in which confidence tier they represent (per
the workbook's own README sheet):

  - "Core Suppliers"        (score >= 3)  322 rows, high-confidence
  - "Extended - Review"     (score 1-2)   354 rows, worth a manual glance
  - "Tier-2 (Non-Trailer)"  (score <= 0)  3246 rows, explicitly NOT
                                          trailer-relevant

This normalizer handles the shared column layout; which sheet(s) to
actually import is a caller decision (see
`scripts/import_automechanika.py`), not this module's — Tier-2 rows
are still real companies with real websites, worth having in the
database for later expansion even though they aren't trailer
suppliers today.

Field mapping decisions
--------------------------
- `website` -> `domain`, run through the same `extract_domain` every
  other normalizer uses, so this participates in the identical
  domain-based dedup matching as a live-scraped record. 47 of 322
  Core Suppliers rows have no website at all -- those still import
  (a company with a name and address is still worth having), just
  without a domain to enrich from until `find-websites` finds one.
- `address` -> stored as-is in `address`, AND `country` is parsed out
  as the last comma-separated segment. Checked against every sample
  row in the real file before trusting this: "...Castleblayney,
  Ireland", "...Ludhiana (Punjab) 141010, India", "...Bessenbach,
  Germany" -- the country is consistently the final segment,
  including when a postal code precedes it in the same segment.
- `matched_product_groups` -> `product_keywords`, split on `;` -- this
  is Automechanika's own structured exhibitor taxonomy (e.g. "Axle
  suspensions", "Fifth-wheel couplings"), which is exactly the shape
  `search_suppliers_full`'s product-text matching already searches
  against. This is real, curated signal, not a free-text guess.
- `description` -> `notes`, kept as reference context for a human
  reviewing a match, never fed into automated scoring (it's marketing
  copy, no different from any other self-reported text).
- `hall_stand` (booth number), `final_score`, `tier`, and `url` (the
  Automechanika exhibitor profile page, NOT the company's own site --
  a real, important distinction from `website`) are NOT mapped to any
  supplier column. They're not lost: `import_static_supplier_list`
  saves every raw row to `raw_source_data` verbatim before
  normalising, so this metadata stays queryable if ever needed, without
  this codebase needing a dedicated column for a one-off import's own
  scoring metadata.
"""

from __future__ import annotations

from typing import Any, Dict, List

from deduplication.domain_utils import extract_domain
from normalizers.base_normalizer import BaseNormalizer

# The three sheet names as they actually appear in the real uploaded
# workbook -- verified directly against the file, not assumed.
SHEET_CORE = "Core Suppliers"
SHEET_EXTENDED = "Extended - Review"
SHEET_TIER2 = "Tier-2 (Non-Trailer)"

# The default import scope: the two trailer-relevant tiers. Tier-2 is
# explicitly "Non-Trailer" per the workbook's own README sheet --
# real companies, worth having eventually, but not what a trailer
# sourcing search should surface by default. Pass include_tier2=True
# to `read_automechanika_workbook` to pull it in too.
DEFAULT_SHEETS = (SHEET_CORE, SHEET_EXTENDED)


def read_automechanika_workbook(
    path: str, *, include_tier2: bool = False
) -> List[Dict[str, Any]]:
    """Reads the real workbook's three sheets (see this module's own
    docstring for their exact shared column layout) and returns one
    raw dict per row, in original column names, completely unmodified
    -- ready to pass straight to
    `pipeline.static_list_import.import_static_supplier_list` alongside
    `AutomechanikaNormaliser()`.

    Deliberately returns raw dicts, not normalised candidates: keeping
    the read and the normalise steps separate matches
    `import_static_supplier_list`'s own two-piece design (raw storage
    happens before normalisation, so nothing is lost to a parsing bug
    in either step).
    """
    import openpyxl

    sheets_to_read = list(DEFAULT_SHEETS)
    if include_tier2:
        sheets_to_read.append(SHEET_TIER2)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    for sheet_name in sheets_to_read:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        sheet_rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(sheet_rows)
        except StopIteration:
            continue
        for row in sheet_rows:
            record = dict(zip(headers, row))
            record["_sheet"] = sheet_name  # preserved in raw_source_data for later reference
            rows.append(record)
    return rows


class AutomechanikaNormaliser(BaseNormalizer):
    source_name = "automechanika_2026"

    def normalise(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        name = self.clean_str(raw_data.get("name"))
        website = self.clean_str(raw_data.get("website"))
        address = self.clean_str(raw_data.get("address"))
        product_groups = self.clean_str(raw_data.get("matched_product_groups"))
        description = self.clean_str(raw_data.get("description"))

        result: Dict[str, Any] = {"canonical_name": name}

        if website:
            domain = extract_domain(website)
            if domain:
                result["domain"] = domain

        if address:
            result["address"] = address
            country = self._parse_country(address)
            if country:
                result["country"] = country

        if product_groups:
            result["product_keywords"] = [
                term.strip() for term in product_groups.split(";") if term.strip()
            ]

        if description:
            result["notes"] = description

        return result

    @staticmethod
    def _parse_country(address: str) -> str:
        """The country is consistently the last comma-separated
        segment in this file's address format -- verified against a
        sample spanning Ireland, Italy, Belgium, India, Germany,
        Taiwan, and Bosnia and Herzegovina before relying on it. Not a
        general-purpose address parser; specific to this export's own
        consistent formatting.
        """
        parts = [p.strip() for p in address.split(",") if p.strip()]
        return parts[-1] if parts else ""
