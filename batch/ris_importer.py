"""
batch/ris_importer.py

One-time import of reverse-image-search evidence (Verification Flag,
Exact Duplicate Domains, Other Matching Domains) from an external
local geocode/Street-View/RIS pipeline's own spreadsheet output --
there is no live RIS pipeline in this codebase, this module only ever
reads a file someone else's tool already produced and writes it into
the DB, the same "source file as one-time input, DB as the durable
home going forward" pattern this codebase already uses for reference
data (see data/source_files/ and its own .gitignore carve-out).

Matching discipline mirrors the earlier ad-hoc enrichment merge this
same spreadsheet was used for: domain first (exact, via
deduplication.domain_utils.extract_domain), falling back to fuzzy
company-name matching (rapidfuzz + deduplication.name_utils.
normalise_company_name, same tool discovery.candidate_validator.py
and verification.uk_company_verification_service.py already use for
the analogous "is this really the same company" question) only when
no domain match exists. Every source row that fails BOTH is reported
back to the caller, never silently dropped -- see
`import_ris_findings`'s own return shape.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from rapidfuzz import fuzz

from deduplication.domain_utils import extract_domain
from deduplication.name_utils import normalise_company_name
from storage.repository import SupplierRepository

# Same bar discovery.candidate_validator's own _NAME_MATCH_THRESHOLD uses
# for the analogous "is this plausibly the same company" fallback check.
_NAME_MATCH_THRESHOLD = 55.0


@dataclass
class RisImportResult:
    matched_by_domain: int = 0
    matched_by_name: int = 0
    unmatched: List[Dict[str, Any]] = field(default_factory=list)  # source rows with no match at all
    imported_supplier_ids: List[int] = field(default_factory=list)


def _read_supplier_audit_rows(xlsx_path: str) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook["Supplier Audit"]
    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def import_ris_findings(xlsx_path: str, repo: Optional[SupplierRepository] = None) -> RisImportResult:
    """Reads the "Supplier Audit" sheet of `xlsx_path` (same shape as
    batch/tracker_exporter.py's build_tracker_workbook produces, plus
    the Verification Flag/Exact Duplicate Domains/Other Matching
    Domains columns merged in by hand for the injection-moulding-100
    audit) and writes those three fields onto each matched supplier's
    row, via update_supplier_fields_with_history for the same audit
    trail every other write in this codebase gets.

    Idempotent to re-run: re-importing the same file just overwrites
    with the same values (update_supplier_fields_with_history only
    logs an actual change, never raises on a no-op re-import).
    """
    repo = repo or SupplierRepository()
    result = RisImportResult()

    for row in _read_supplier_audit_rows(xlsx_path):
        name = row.get("Supplier Name") or ""
        website = row.get("Website") or ""
        verification_flag = row.get("Verification Flag")
        exact_dup = row.get("Exact Duplicate Domains")
        other_match = row.get("Other Matching Domains")

        domain = extract_domain(website)
        supplier = repo.find_by_domain(domain) if domain else None
        matched_via = "domain" if supplier else None

        if supplier is None and name.strip():
            supplier, _score = _fuzzy_match_by_name(repo, name)
            if supplier is not None:
                matched_via = "name"

        if supplier is None:
            result.unmatched.append({"Supplier Name": name, "Website": website})
            continue

        if matched_via == "domain":
            result.matched_by_domain += 1
        else:
            result.matched_by_name += 1

        from datetime import datetime, timezone

        repo.update_supplier_fields_with_history(
            supplier["id"],
            {
                "ris_verification_flag": verification_flag,
                "ris_exact_duplicate_domains": exact_dup,
                "ris_other_matching_domains": other_match,
                "ris_imported_at": datetime.now(timezone.utc).isoformat(),
            },
            changed_by="ris_importer",
            change_reason=f"one-time RIS evidence import from {xlsx_path}",
        )
        result.imported_supplier_ids.append(supplier["id"])

    return result


def _fuzzy_match_by_name(repo: SupplierRepository, name: str) -> tuple:
    """Deliberately simple/slow (full scan, no index) -- this module
    only ever runs as a one-time, human-invoked import over ~100 rows,
    never in a hot path, so a full-catalogue scan is the honest
    tradeoff against adding new repository/index machinery for a
    single use site. Reuses repo.find_by_country("") -- already this
    codebase's own established "every supplier, for a fuzzy-match
    candidate pool" call (see that method's own docstring: "Used by
    the fuzzy name matcher"), not a new method. Returns
    (best_supplier_or_None, best_score)."""
    normalised_target = normalise_company_name(name)
    best_supplier, best_score = None, 0.0
    for supplier in repo.find_by_country(""):
        score = fuzz.ratio(normalised_target, normalise_company_name(supplier.get("canonical_name") or ""))
        if score > best_score:
            best_supplier, best_score = supplier, score
    if best_score >= _NAME_MATCH_THRESHOLD:
        return best_supplier, best_score
    return None, best_score
