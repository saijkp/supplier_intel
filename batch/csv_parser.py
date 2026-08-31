"""
batch/csv_parser.py

Parses an uploaded CSV into per-row candidates for batch/batch_service.py,
fuzzy-matching messy real-world column headers ("Company", "Website URL",
"URL", "Company Name") onto the two fields batch enrichment actually
needs -- company_name and website. Every original column is preserved
verbatim per row (for export pass-through, see batch/csv_exporter.py) --
this module only detects which columns to READ FROM, it never drops or
renames anything in the source data.

Deliberately does no row classification (needs_url/needs_name/ready) or
enrichment decisions -- that's batch/batch_service.py's job. This module's
only job is "what are the rows, and which columns are company_name/
website," so it can be tested and reasoned about in isolation.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from rapidfuzz import fuzz

# Canonical column aliases, checked via fuzzy match (not exact) against
# each real header -- real spreadsheets vary: "Company", "Company Name",
# "Business Name", "Supplier"; "Website", "URL", "Website URL", "Site".
_COMPANY_NAME_ALIASES: tuple = (
    "company name", "company", "business name", "supplier", "supplier name",
    "organisation", "organization", "org name", "vendor", "vendor name",
)
_WEBSITE_ALIASES: tuple = (
    "website", "website url", "url", "site", "web site", "domain",
    "company website", "web address", "homepage", "link",
)
# Trusted directly into suppliers.country when present -- this is the
# uploader's OWN data, not a scraped extraction, same status as
# _COMPANY_NAME_ALIASES already has for canonical_name (see
# batch_service.py's _resolve_named_row/_resolve_placeholder_row).
_COUNTRY_ALIASES: tuple = ("country", "nation")

# rapidfuzz.fuzz.ratio, 0-100 -- checked against real messy header
# samples ("Company Name", "company", "Business Name ", "COMPANY") in
# tests/test_csv_parser.py before settling on this threshold.
_FUZZY_MATCH_THRESHOLD = 70.0


@dataclass
class ParsedRow:
    row_index: int
    original_columns: Dict[str, str]      # every column from this CSV row, verbatim (whitespace-stripped)
    company_name: Optional[str]           # detected company name, stripped, or None if no column matched / cell empty
    website: Optional[str]                # detected website, stripped, or None if no column matched / cell empty
    country: Optional[str] = None         # detected country, stripped, or None if no column matched / cell empty


@dataclass
class ParseResult:
    rows: List[ParsedRow] = field(default_factory=list)
    company_name_column: Optional[str] = None    # which real header was detected as company_name, or None if none matched
    website_column: Optional[str] = None          # which real header was detected as website, or None if none matched
    country_column: Optional[str] = None          # which real header was detected as country, or None if none matched
    # row_index of any row whose (company_name, website) pair repeats an
    # earlier row -- informational only; batch_service.py decides what
    # to do with it (e.g. skip a redundant real enrichment call and
    # reuse the first occurrence's resolved supplier), this module never
    # drops a row on its own account.
    duplicate_row_indices: List[int] = field(default_factory=list)


def _normalise_header(header: str) -> str:
    return header.strip().lower().replace("_", " ").replace("-", " ")


def _best_column_match(headers: List[str], aliases: tuple) -> Optional[str]:
    """The real header whose normalised text best fuzzy-matches any
    alias, or None if nothing clears _FUZZY_MATCH_THRESHOLD. Ties keep
    the first (leftmost) header, matching how a human skimming column
    order would naturally pick."""
    best_header: Optional[str] = None
    best_score = 0.0
    for header in headers:
        if not header:
            continue
        normalised = _normalise_header(header)
        for alias in aliases:
            score = fuzz.ratio(normalised, alias)
            if score > best_score:
                best_score = score
                best_header = header
    return best_header if best_score >= _FUZZY_MATCH_THRESHOLD else None


def _build_parse_result(headers: List[str], row_dicts: List[Dict[str, str]]) -> ParseResult:
    """Shared by parse_csv and parse_xlsx once each has reduced its own
    file format down to (headers, one stripped-string dict per row) --
    fuzzy column detection and per-row extraction/dedup only need to be
    written, and tested, once."""
    result = ParseResult()
    if not headers:
        return result

    company_col = _best_column_match(headers, _COMPANY_NAME_ALIASES)
    website_col = _best_column_match(headers, _WEBSITE_ALIASES)
    country_col = _best_column_match(headers, _COUNTRY_ALIASES)
    result.company_name_column = company_col
    result.website_column = website_col
    result.country_column = country_col

    seen_keys: set = set()
    for i, original in enumerate(row_dicts):
        company_name = (original.get(company_col) or "").strip() or None if company_col else None
        website = (original.get(website_col) or "").strip() or None if website_col else None
        country = (original.get(country_col) or "").strip() or None if country_col else None

        key = (company_name or "", website or "")
        if key != ("", "") and key in seen_keys:
            result.duplicate_row_indices.append(i)
        else:
            seen_keys.add(key)

        result.rows.append(ParsedRow(
            row_index=i, original_columns=original,
            company_name=company_name, website=website, country=country,
        ))

    return result


def parse_csv(file_content: bytes) -> ParseResult:
    """Never raises for ordinary messy input (empty file, no recognisable
    header, malformed rows) -- returns an empty/partial ParseResult
    instead, matching every other source's "return partial data, don't
    lose the batch to a parsing bug" discipline already established in
    this codebase (see normalizers/base_normalizer.py)."""
    if not file_content:
        return ParseResult()

    try:
        text = file_content.decode("utf-8-sig")  # -sig strips a BOM, common in Excel-exported CSVs
    except UnicodeDecodeError:
        text = file_content.decode("utf-8", errors="replace")

    try:
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
    except csv.Error:
        return ParseResult()
    if not headers:
        return ParseResult()

    row_dicts = [
        {k: (v or "").strip() for k, v in raw_row.items() if k is not None}
        for raw_row in reader
    ]
    return _build_parse_result(headers, row_dicts)


def parse_xlsx(file_content: bytes) -> ParseResult:
    """Same fuzzy header-detection/per-row extraction as parse_csv, just
    reading the first worksheet of an uploaded .xlsx workbook instead of
    decoding CSV text -- POST /batch/upload's Find Suppliers bulk-upload
    flow accepts either now (see parse_batch_upload_file). Never raises
    for a malformed/empty/password-protected workbook -- returns an
    empty ParseResult, matching parse_csv's own discipline."""
    if not file_content:
        return ParseResult()

    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            return ParseResult()
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
    except Exception:  # noqa: BLE001 -- any corrupt/unreadable workbook degrades to empty, never raises
        return ParseResult()

    if not header_row:
        return ParseResult()

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    row_dicts = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None for v in raw_row):
            continue  # openpyxl can yield a fully-blank trailing row
        row_dicts.append({
            headers[idx]: (str(v).strip() if v is not None else "")
            for idx, v in enumerate(raw_row) if idx < len(headers)
        })
    return _build_parse_result(headers, row_dicts)


def parse_batch_upload_file(file_content: bytes, filename: Optional[str] = None) -> ParseResult:
    """The single entry point POST /batch/upload uses -- dispatches to
    parse_xlsx or parse_csv by the uploaded file's extension so the
    caller doesn't need to know which format it got. Defaults to CSV
    when the filename is missing or unrecognised, matching this
    endpoint's original CSV-only behaviour exactly. Deliberately does
    NOT route legacy .xls (openpyxl only reads .xlsx/.xlsm -- adding
    real .xls support would need a second library, not requested)."""
    ext = filename.rsplit(".", 1)[-1].lower() if filename and "." in filename else ""
    if ext == "xlsx":
        return parse_xlsx(file_content)
    return parse_csv(file_content)
