"""
batch/tracker_exporter.py

Group 3 of the procurement-tracker feature set: exports a list of
already-confirmed candidate suppliers into the buyer's own tracker CSV
format -- exact column order/names matching their real "Injection
Moulding" tab (No. through Date Reviewed) so rows paste in without
reformatting or shifting any of the tracker's own columns.

Columns A-D and Qualified are the literal text "Pending" until a human
sets a real value via the Audit tab's Stage 2 verdict controls (see
storage/repository.py's audit_verdicts table and
SupplierRepository.upsert_audit_verdict) -- "Pending" is an explicit
placeholder, not a verdict: it exists so a blank cell is never
ambiguous between "not yet reviewed" and "forgot to check." This
module (and everything upstream of it) still never COMPUTES or INFERS
a Pass/Fail/Yes/No itself -- audit_verdicts only ever stores a direct,
unmodified echo of what a human selected in the UI, and
_verdict_or_pending below reads that back verbatim or falls back to
"Pending" if nothing's been set. See PASTE_RANGE_COLUMNS's own
comment.

Evidence/helper columns (a Website Note, per-field source URLs,
candidate facility photos, a Street View link, a LinkedIn search link,
certification evidence, catalogue-depth evidence (customer-logos
section / named case studies / specific process-machine detail, blank
unless main.py extract-catalogue-depth has run -- see
verification/catalogue_depth_service.py, including a real caveat
there about its source_url being a local file path, not a web link),
the discovery pipeline's own gate-pass reason text, Companies House
match status (blank unless main.py verify-uk-company has run -- see
standing rule 9 in CLAUDE.md), D-search snippets, Checks Remaining,
Sort Key) are appended AFTER Date Reviewed -- outside the paste range, so the
main block still pastes in clean while the evidence needed to audit
each value stays in the same file, one click away. Street View/
LinkedIn links are both free Google search-links, never an automated
match judgment -- deliberately no proxy infrastructure and no reverse
image search here (out of scope given the deadline: most dead fetches
are DNS failures a proxy wouldn't fix, and neither of us has a working
reverse-image-search tool right now).

Website Note: flags the case where SupplierMatcher's fuzzy-name merge
kept an old/dead stored domain on a supplier record while collection
actually crawled a DIFFERENT, real, working domain (found via real
gap-analysis recoveries -- a company's canonical_name matched an
existing record closely enough to auto-merge, but merge_into_golden
only fills EMPTY fields, so the stale domain was never overwritten
even though the new site was what actually got scraped). Detected by
comparing the domain embedded in the address/factory_location
field_provenance source_url against suppliers.domain -- blank when
they agree (the overwhelming majority of rows) or when there's no
provenance to compare against at all.

A second export, the removed-candidates list, mirrors the tracker's
own "Removed Candidates" tab (Company Name, Website, Reason) -- every
supplier a human has flagged=1 excluded (see storage/repository.py's
flagged/flag_reason fields, wired into search_suppliers/
search_suppliers_full/find_discovered_suppliers so an excluded record
can't silently resurface as a candidate later) among a given set of
supplier ids.
"""

from __future__ import annotations

import csv
import io
import urllib.parse
from typing import Any, Dict, List, Optional, Tuple

from deduplication.domain_utils import extract_domain
from storage.repository import SupplierRepository

# Exact column order/names of the buyer's real tracker "Injection
# Moulding" tab, confirmed against their own exported CSV. A-D/
# Qualified/Notes/Date Reviewed reflect real audit_verdicts rows when
# they exist (see _verdict_or_pending), "Pending"/blank otherwise --
# see module docstring.
PASTE_RANGE_COLUMNS: Tuple[str, ...] = (
    "No.", "Supplier Name", "Website", "Country", "Address", "Phone", "Email",
    "Factory Location",
    "A - Website Deep-Dive", "B - Certifications", "C - Factory Authenticity",
    "D - Reviews & Ratings", "Qualified", "Notes / Difficulties", "Date Reviewed",
)

# The literal placeholder written into A-D/Qualified/Notes/Date Reviewed
# on export when no human has reviewed that criterion yet -- an explicit
# "not yet reviewed" marker, never a verdict (see module docstring for
# why a blank cell alone is ambiguous here). Once a real audit_verdicts
# row exists (see storage.repository.SupplierRepository
# .upsert_audit_verdict, written only from a human's own Audit tab
# selection), _verdict_or_pending below returns that real value instead.
_PENDING_PLACEHOLDER = "Pending"


def _verdict_or_pending(verdicts: Dict[str, Dict[str, Any]], criterion: str) -> str:
    """The real value a human has set for `criterion` (A/B/C/D/
    Qualified), or _PENDING_PLACEHOLDER if nothing has been set yet.
    `verdicts` is repo.get_audit_verdicts(supplier_id)'s own return
    shape -- a criterion with no row simply isn't a key."""
    row = verdicts.get(criterion)
    if row is None or not row.get("value"):
        return _PENDING_PLACEHOLDER
    return row["value"]

# Appended after the paste range -- never part of what gets pasted
# into the tracker itself. Phone/Email only have row-level
# (contact_source_pages), not per-value, source tracking -- see
# storage/database.py's contact_source_pages comment -- so those two
# columns show the page(s) checked for contact info, not a single
# precise per-value link, unlike Address/Factory Location which do
# have exact per-value provenance (field_provenance).
EVIDENCE_COLUMNS: Tuple[str, ...] = (
    "Website Note", "Address Source URL", "Phone Source Page(s)", "Email Source Page(s)",
    "Factory Location Source URL",
    "Candidate Facility Photo URLs", "Street View Link", "LinkedIn Search Link",
    "Certifications Claimed (source + evidence)",
    "Catalogue Depth Signals (source + evidence)",
    "Discovery Validation Reason", "Companies House Status",
    "D-Search: Scam", "D-Search: Review", "D-Search: Factory Tour",
    "Checks Remaining", "Sort Key (helper)",
)

REMOVED_CANDIDATES_COLUMNS: Tuple[str, ...] = ("Company Name", "Website", "Reason")

# Every field a fresh export leaves entirely to the buyer -- see
# module docstring. Always all four: this export only ever produces
# brand-new candidate rows, which never have any of A-D pre-filled.
_CHECKS_REMAINING_ON_A_FRESH_EXPORT = "A, B, C, D"

# The fields evidence_score checks -- one point each, so Sort Key
# (helper) ranges 0-4. Used to rank rows by how much evidence is
# already on file (most review-ready first), never as a Qualified
# judgment -- it says nothing about whether the evidence is GOOD, only
# whether it exists yet.
_EVIDENCE_SCORE_FIELDS: Tuple[str, ...] = ("address", "primary_phone", "primary_email", "factory_location")


def evidence_score(supplier: Dict[str, Any]) -> int:
    """Public (not module-private) since api/app.py's Audit endpoints
    also sort a supplier list by this -- same "list confirmed
    suppliers by evidence completeness" question build_tracker_export's
    own row-sort already answers, reused rather than reimplemented."""
    return sum(1 for field in _EVIDENCE_SCORE_FIELDS if supplier.get(field))


def _latest_provenance_source_url(repo: SupplierRepository, supplier_id: int, field_name: str) -> str:
    entries = repo.get_field_provenance(supplier_id, field_name=field_name)
    return (entries[-1].get("source_url") or "") if entries else ""


def _bare_domain(domain: str) -> str:
    d = domain.lower()
    return d[4:] if d.startswith("www.") else d


def _website_note(repo: SupplierRepository, supplier_id: int, stored_domain: str) -> str:
    """See module docstring's own "Website Note" section. Checks
    address then factory_location provenance (the two fields most
    likely to have been extracted from a real, freshly-crawled page)
    for a source_url whose domain disagrees with what's actually
    stored on the supplier record. Blank when they agree or when
    there's nothing to compare against."""
    if not stored_domain:
        return ""
    for field_name in ("address", "factory_location"):
        entries = repo.get_field_provenance(supplier_id, field_name=field_name)
        if not entries:
            continue
        crawled_domain = extract_domain(entries[-1].get("source_url") or "")
        if not crawled_domain:
            continue
        if _bare_domain(crawled_domain) == _bare_domain(stored_domain):
            return ""
        return (
            f"Resolved via {crawled_domain} during collection -- stored/registered domain "
            f"({stored_domain}) may be dead or different. Evidence above (address/factory "
            f"location source URLs) is from the working site; the Website column reflects "
            f"the supplier record's own stored domain, not necessarily a live link."
        )
    return ""


def _phone_source_pages(repo: SupplierRepository, supplier: Dict[str, Any]) -> str:
    # No per-phone source_url match attempted beyond the exact primary_phone
    # string -- supplier_phone_numbers.source_url exists, but
    # contact_source_pages is the honest fallback for what's actually
    # reliably available at export time (see EVIDENCE_COLUMNS comment).
    return "; ".join(supplier.get("contact_source_pages") or [])


def _street_view_link(address: str) -> str:
    """Free Google Maps search-link (no geocoding/Street-View-Static
    API dependency) -- one click to the address's approximate map
    location, not an automated match judgment. Empty when there's no
    address to link."""
    if not address:
        return ""
    return f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(address)}"


def _linkedin_search_link(company_name: str) -> str:
    """Same free-search-link pattern as _street_view_link, no LinkedIn
    API/scraping involved -- a site:linkedin.com Google search for the
    company name, one click for the buyer to manually check for a real
    company page. Deliberately not an automated match/verdict, and
    deliberately not proxy infrastructure or reverse image search (out
    of scope for now -- most dead fetches are DNS failures a proxy
    wouldn't fix, and neither of us has a working reverse-image-search
    tool). Empty when there's no name to search for."""
    if not company_name:
        return ""
    query = f'site:linkedin.com "{company_name}"'
    return f"https://www.google.com/search?q={urllib.parse.quote(query)}"


def _marketplace_search_link(company_name: str, site: str) -> str:
    """Same free-search-link pattern as _linkedin_search_link -- a
    site:<marketplace> Google search for the company name. Presence on
    Alibaba/Made-in-China is a neutral signal on its own (real
    manufacturers list there too); the buyer reads the result and
    judges it themselves -- e.g. whether a supplier is ONLY findable
    there is the kind of thing worth noticing, not something this
    function decides. Empty when there's no name to search for."""
    if not company_name:
        return ""
    query = f'site:{site} "{company_name}"'
    return f"https://www.google.com/search?q={urllib.parse.quote(query)}"


def _trustpilot_search_link(company_name: str) -> str:
    """Trustpilot's own search page, not a Google site: search --
    confirmed live that trustpilot.com/search?query= is a real,
    working query parameter (no CAPTCHA, no JS-only form). Empty when
    there's no name to search for."""
    if not company_name:
        return ""
    return f"https://www.trustpilot.com/search?query={urllib.parse.quote(company_name)}"


# Plain links to the correct registry search PAGE per certification standard --
# NOT deep links to a specific certificate result. Confirmed live (browser
# session) that neither registry supports a bookmarkable/parameterized direct
# link to a certificate: IAF CertSearch's search box doesn't change the URL
# and is gated by a reCAPTCHA on every query; IATF Customer Portal's search
# returns a real result with no CAPTCHA, but the result URL never encodes the
# certificate number (confirmed via window.location.href) -- it's a client-
# side form submission, not a GET with a query string. So this is a link to
# the right search page only -- the buyer pastes the certificate number in
# themselves once there, same as every other link in this evidence bundle.
_CERTIFICATION_REGISTRY_LINKS = {
    "iso 9001": "https://www.iafcertsearch.org/",
    "iso 14001": "https://www.iafcertsearch.org/",
    "iso 13485": "https://www.iafcertsearch.org/",
    "iatf 16949": "https://iatf-customerportal.org/",
}


def _certification_registry_link(canonical_term: Optional[str]) -> str:
    return _CERTIFICATION_REGISTRY_LINKS.get(canonical_term or "", "")


def _certifications_claimed(repo: SupplierRepository, supplier_id: int) -> str:
    """Export-only surfacing of certification claims the older,
    separate capability_extractor.py stage already found (category=
    'standard' in supplier_capabilities) -- no new extraction here.
    Blank for a supplier that stage has never run against."""
    parts = []
    for cap in repo.get_capabilities(supplier_id):
        if cap.get("category") != "standard":
            continue
        term = cap.get("canonical_term") or cap.get("reported_term") or ""
        source = cap.get("source_url") or ""
        evidence = (cap.get("evidence") or "").strip()
        parts.append(f'{term} [{source}] "{evidence}"')
    return " | ".join(parts)


_CATALOGUE_SIGNAL_LABELS = {
    "customer_logos_section": "Customer logos",
    "named_case_studies": "Named case study",
    "specific_process_detail": "Specific process/machine detail",
    "self_described_role": "Self-described role",
}


def _catalogue_depth_signals(repo: SupplierRepository, supplier_id: int) -> str:
    """Export-only surfacing of catalogue-depth evidence the separate
    catalogue_depth_extractor.py stage already found (customer-logos
    section / named case studies / specific process-machine detail) --
    no new extraction here, same pattern as _certifications_claimed.
    Blank for a supplier that stage has never run against. Note:
    `source_url` here is a LOCAL ARTIFACT FILE PATH, not a web URL --
    see verification/catalogue_depth_service.py's own docstring for
    why (SiteCollector never persists a per-page URL mapping after a
    collection run ends, and guessing one from the lossy filename slug
    would present an inferred value as a stated fact)."""
    parts = []
    for sig in repo.get_catalogue_signals(supplier_id):
        label = _CATALOGUE_SIGNAL_LABELS.get(sig.get("signal_type"), sig.get("signal_type") or "")
        source = sig.get("source_url") or ""
        evidence = (sig.get("evidence") or "").strip()
        parts.append(f'{label} [{source}] "{evidence}"')
    return " | ".join(parts)


def _discovery_validation_reason(repo: SupplierRepository, supplier_id: int) -> str:
    """Export-only surfacing of discovery.candidate_validator's own
    gate reason text for whichever raw_source_data row most recently
    fed this supplier (e.g. "validated: name corroborated (score=100),
    product term found on page" -- gate 6's own wording, see that
    module's REASON_* constants) -- evidence of WHY discovery accepted
    the candidate, never a re-judgment of it. Blank for a supplier this
    export-adjacent stage never discovered (e.g. manually batch-
    uploaded)."""
    rows = repo.get_raw_by_golden_record(supplier_id)
    for row in rows:
        raw_json = row.get("raw_json")
        if isinstance(raw_json, dict) and raw_json.get("reason"):
            return raw_json["reason"]
    return ""


def _companies_house_status(supplier: Dict[str, Any]) -> str:
    """Export-only surfacing of verification.uk_company_verification_service's
    own match outcome -- never re-derived or re-judged here, same
    "surface the evidence, not a verdict" discipline as every other
    EVIDENCE_COLUMNS field (see module docstring / standing rule 2).
    "no_clear_match" is deliberately not itself an exclusion signal
    (see that service's own docstring) -- surfaced exactly as found,
    the buyer's own manual check either way. Blank for a supplier
    main.py verify-uk-company has never run against."""
    status = supplier.get("companies_house_match_status")
    if not status:
        return ""
    confidence = supplier.get("companies_house_match_confidence")
    parts = [f"match_status={status}"]
    if confidence is not None:
        parts.append(f"confidence={confidence}")
    company_status = supplier.get("companies_house_status")
    if company_status:
        parts.append(f"company_status={company_status}")
    number = supplier.get("companies_house_number")
    if number:
        parts.append(f"company_number={number}")
    office = supplier.get("companies_house_registered_office")
    if office:
        parts.append(f"registered_office={office}")
    return "; ".join(parts)


def _reputation_snippets(repo: SupplierRepository, supplier_id: int, query_type: str) -> str:
    """Export-only surfacing of Group 2's opt-in D-search results.
    Blank unless --search-reputation was actually run for this
    supplier -- never a Clean/Flagged judgment, just the raw
    title/snippet/link each query returned."""
    parts = []
    for s in repo.get_reputation_snippets(supplier_id, query_type=query_type):
        title = s.get("title") or ""
        snippet = s.get("snippet") or ""
        link = s.get("link") or ""
        parts.append(f"{title}: {snippet} ({link})")
    return " | ".join(parts)


def build_tracker_export(supplier_ids: List[int], repo: Optional[SupplierRepository] = None) -> str:
    """Returns CSV text (not bytes) -- caller decides encoding/response
    headers. A supplier_id with no matching row is silently skipped
    (never raises) -- one stale/deleted id must never break the whole
    export.

    Row order: Sort Key (helper) descending -- most evidence-complete
    candidates first -- ties broken alphabetically by name for a
    stable, predictable order across repeated exports of the same set."""
    repo = repo or SupplierRepository()

    scored: List[Tuple[int, str, int, Dict[str, Any]]] = []
    for supplier_id in supplier_ids:
        supplier = repo.get_supplier(supplier_id)
        if supplier is None:
            continue
        scored.append((
            evidence_score(supplier), (supplier.get("canonical_name") or "").lower(), supplier_id, supplier,
        ))
    scored.sort(key=lambda t: (-t[0], t[1]))

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([*PASTE_RANGE_COLUMNS, *EVIDENCE_COLUMNS])

    for i, (score, _, supplier_id, supplier) in enumerate(scored, start=1):
        address = supplier.get("address") or ""
        phone = supplier.get("primary_phone") or ""
        email = supplier.get("primary_email") or ""
        factory_location = supplier.get("factory_location") or ""
        verdicts = repo.get_audit_verdicts(supplier_id)

        paste_row = [
            i, supplier.get("canonical_name") or "", supplier.get("domain") or "",
            supplier.get("country") or "", address, phone, email, factory_location,
            # A, B, C, D, Qualified -- the real value a human set via the Audit
            # tab, "Pending" if nothing's been set yet. See module docstring --
            # this module still never computes/infers a verdict itself.
            _verdict_or_pending(verdicts, "A"), _verdict_or_pending(verdicts, "B"),
            _verdict_or_pending(verdicts, "C"), _verdict_or_pending(verdicts, "D"),
            _verdict_or_pending(verdicts, "Qualified"),
            (verdicts.get("Notes") or {}).get("notes") or "",
            repo.get_audit_review_date(supplier_id) or "",
        ]

        evidence_row = [
            _website_note(repo, supplier_id, supplier.get("domain") or ""),
            _latest_provenance_source_url(repo, supplier_id, "address"),
            _phone_source_pages(repo, supplier),
            "; ".join(supplier.get("contact_source_pages") or []),
            _latest_provenance_source_url(repo, supplier_id, "factory_location"),
            "; ".join(supplier.get("candidate_facility_photo_urls") or []),
            _street_view_link(address),
            _linkedin_search_link(supplier.get("canonical_name") or ""),
            _certifications_claimed(repo, supplier_id),
            _catalogue_depth_signals(repo, supplier_id),
            _discovery_validation_reason(repo, supplier_id),
            _companies_house_status(supplier),
            _reputation_snippets(repo, supplier_id, "scam"),
            _reputation_snippets(repo, supplier_id, "review"),
            _reputation_snippets(repo, supplier_id, "factory_tour"),
            _CHECKS_REMAINING_ON_A_FRESH_EXPORT,
            score,
        ]
        writer.writerow([*paste_row, *evidence_row])

    return output.getvalue()


def build_removed_candidates_export(
    supplier_ids: List[int], repo: Optional[SupplierRepository] = None,
) -> str:
    """Every supplier among `supplier_ids` that's flagged=1 excluded --
    mirrors the tracker's own "Removed Candidates" tab. Scoped to the
    ids passed in (the current export's own candidate universe), not
    every flagged supplier system-wide -- a company excluded from a
    DIFFERENT product category's sourcing run has no place on THIS
    tracker's removed-candidates list."""
    repo = repo or SupplierRepository()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(REMOVED_CANDIDATES_COLUMNS)

    for supplier_id in supplier_ids:
        supplier = repo.get_supplier(supplier_id)
        if supplier is None or not supplier.get("flagged"):
            continue
        domain = supplier.get("domain") or ""
        website = f"https://{domain}" if domain else ""
        writer.writerow([supplier.get("canonical_name") or "", website, supplier.get("flag_reason") or ""])

    return output.getvalue()


def _write_csv_text_to_sheet(sheet: Any, csv_text: str) -> None:
    """Parses `csv_text` (already-quoted CSV, e.g. build_tracker_export's
    own output) back into rows and writes them into an openpyxl sheet --
    round-trips cleanly since Python's csv module reads back exactly
    what it wrote, including any embedded commas/quotes/newlines in an
    address or evidence field. Bolds + freezes the header row and
    auto-sizes columns, same look as the existing xlsx exports
    (reports.generator.suppliers_to_excel_bytes). A header-only
    csv_text (e.g. build_tracker_export([], repo)) still gets the
    header formatting, just no data rows below it."""
    from openpyxl.styles import Font

    all_rows = list(csv.reader(io.StringIO(csv_text)))
    for row in all_rows:
        sheet.append(row)
    if not all_rows:
        return
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for i, value in enumerate(all_rows[0], start=1):
        width = max(len(str(value)), 10)
        sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = min(width + 6, 55)


def build_tracker_workbook(
    confirmed_supplier_ids: List[int], removed_universe_supplier_ids: List[int],
    repo: Optional[SupplierRepository] = None,
) -> bytes:
    """Same three-tab shape as the buyer's own reference tracker file
    (Supplier Audit / Qualified / Removed Candidates), built ON TOP of
    build_tracker_export/build_removed_candidates_export above rather
    than duplicating their row-construction logic -- the .xlsx output
    can never drift from what the CSV exports produce, since it's
    quite literally their own output re-parsed into sheets.

    Supplier Audit and Qualified share identical columns (the same
    paste range + evidence columns build_tracker_export always writes)
    -- Qualified starts as a header row only, zero data rows, meant for
    the buyer to move rows into by hand once actually reviewed; nothing
    in this codebase ever writes a row into it itself, matching the
    same "never write a verdict, that's the buyer's manual call"
    discipline documented at the top of this module. Removed Candidates
    mirrors build_removed_candidates_export exactly: Company Name/
    Website/Reason for every flagged=1 supplier among
    `removed_universe_supplier_ids`.

    Returns raw .xlsx bytes (same convention as
    reports.generator.suppliers_to_excel_bytes) -- caller decides the
    output path/write mode."""
    from openpyxl import Workbook

    repo = repo or SupplierRepository()

    workbook = Workbook()
    audit_sheet = workbook.active
    audit_sheet.title = "Supplier Audit"
    _write_csv_text_to_sheet(audit_sheet, build_tracker_export(confirmed_supplier_ids, repo))

    qualified_sheet = workbook.create_sheet("Qualified")
    _write_csv_text_to_sheet(qualified_sheet, build_tracker_export([], repo))

    removed_sheet = workbook.create_sheet("Removed Candidates")
    _write_csv_text_to_sheet(removed_sheet, build_removed_candidates_export(removed_universe_supplier_ids, repo))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_simple_tracker_workbook(supplier_ids: List[int], repo: Optional[SupplierRepository] = None) -> bytes:
    """A single-sheet .xlsx wrapper around build_tracker_export -- for
    an arbitrary supplier_id list with no category/roster/removed-
    candidates concept of its own (e.g. GET /pipeline/jobs/{id}/export.xlsx,
    a Find Suppliers run's own results). build_tracker_workbook above is
    the category-tracker-specific three-tab shape (Supplier Audit /
    Qualified / Removed Candidates); this is the one-tab equivalent for
    any other caller that just wants "these supplier_ids, tracker
    format, as a spreadsheet" without the category machinery."""
    from openpyxl import Workbook

    repo = repo or SupplierRepository()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Suppliers"
    _write_csv_text_to_sheet(sheet, build_tracker_export(supplier_ids, repo))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_supplier_evidence_bundle(
    supplier_id: int, repo: Optional[SupplierRepository] = None,
) -> Optional[Dict[str, Any]]:
    """One supplier's complete evidence picture as structured JSON --
    built for api/app.py's Audit endpoints (GET /audit/suppliers/{id}),
    reusing every one of this module's own evidence helpers verbatim
    (_website_note, _latest_provenance_source_url, _phone_source_pages,
    _street_view_link, _linkedin_search_link, evidence_score,
    _discovery_validation_reason) rather than re-deriving any of them.
    Multi-item evidence (certifications, catalogue-depth signals,
    reputation snippets) is read from the same repository methods
    build_tracker_export itself calls, just returned as structured
    lists here instead of the single flattened "term [url] "evidence""
    string a CSV cell needs -- same underlying data, better shape for
    a web page to render.

    Returns None if no supplier exists with this id -- the caller
    (an API route) turns that into a 404, this module has no HTTP
    concept of its own.

    Purely a read -- assembles and returns already-stored evidence,
    never calls an LLM and never writes anything. `verdicts` (Stage 2)
    is the one exception to "evidence, not verdict": it echoes back
    whatever a human has already selected via PUT
    /audit/suppliers/{id}/verdicts/{criterion} (storage.repository
    .SupplierRepository.upsert_audit_verdict), never a value this
    function or anything upstream of it computes itself -- same
    standing-rule-2 discipline as every other A-D/Qualified column in
    this codebase, just now backed by a real human's own input instead
    of always being "Pending"."""
    repo = repo or SupplierRepository()
    supplier = repo.get_supplier(supplier_id)
    if supplier is None:
        return None

    address = supplier.get("address") or ""

    return {
        "supplier_id": supplier_id,
        "canonical_name": supplier.get("canonical_name") or "",
        "domain": supplier.get("domain") or "",
        "country": supplier.get("country") or "",
        "evidence_score": evidence_score(supplier),
        "flagged": bool(supplier.get("flagged")),
        "flag_reason": supplier.get("flag_reason") or "",
        "contact": {
            "address": address,
            "phone": supplier.get("primary_phone") or "",
            "email": supplier.get("primary_email") or "",
            "factory_location": supplier.get("factory_location") or "",
        },
        "website_note": _website_note(repo, supplier_id, supplier.get("domain") or ""),
        "source_urls": {
            "address": _latest_provenance_source_url(repo, supplier_id, "address"),
            "factory_location": _latest_provenance_source_url(repo, supplier_id, "factory_location"),
            "phone_source_pages": _phone_source_pages(repo, supplier),
            "email_source_pages": list(supplier.get("contact_source_pages") or []),
        },
        "candidate_facility_photo_urls": list(supplier.get("candidate_facility_photo_urls") or []),
        "street_view_link": _street_view_link(address),
        "linkedin_search_link": _linkedin_search_link(supplier.get("canonical_name") or ""),
        "alibaba_search_link": _marketplace_search_link(supplier.get("canonical_name") or "", "alibaba.com"),
        "made_in_china_search_link": _marketplace_search_link(supplier.get("canonical_name") or "", "made-in-china.com"),
        "trustpilot_search_link": _trustpilot_search_link(supplier.get("canonical_name") or ""),
        "companies_house": {
            "match_status": supplier.get("companies_house_match_status"),
            "match_confidence": supplier.get("companies_house_match_confidence"),
            "status": supplier.get("companies_house_status"),
            "number": supplier.get("companies_house_number"),
            "registered_office": supplier.get("companies_house_registered_office"),
            "incorporated_at": supplier.get("companies_house_incorporated_at"),
            "checked_at": supplier.get("companies_house_checked_at"),
        },
        "certifications_claimed": [
            {
                "term": cap.get("canonical_term") or cap.get("reported_term"),
                "source_url": cap.get("source_url"),
                "evidence": cap.get("evidence"),
                "registry_link": _certification_registry_link(cap.get("canonical_term")),
            }
            for cap in repo.get_capabilities(supplier_id) if cap.get("category") == "standard"
        ],
        "capability_extraction_status": supplier.get("capability_extraction_status"),
        "catalogue_depth_signals": [
            {"signal_type": s.get("signal_type"), "evidence": s.get("evidence"), "source_url": s.get("source_url")}
            for s in repo.get_catalogue_signals(supplier_id)
        ],
        "reverse_image_search": {
            "verification_flag": supplier.get("ris_verification_flag"),
            "exact_duplicate_domains": supplier.get("ris_exact_duplicate_domains"),
            "other_matching_domains": supplier.get("ris_other_matching_domains"),
            "imported_at": supplier.get("ris_imported_at"),
        },
        "reputation_snippets": {
            query_type: [
                {"title": s.get("title"), "snippet": s.get("snippet"), "link": s.get("link")}
                for s in repo.get_reputation_snippets(supplier_id, query_type=query_type)
            ]
            for query_type in ("scam", "review", "factory_tour")
        },
        "reputation_search_attempted_at": supplier.get("reputation_search_attempted_at"),
        "discovery_validation_reason": _discovery_validation_reason(repo, supplier_id),
        "verdicts": repo.get_audit_verdicts(supplier_id),
        # Only ever populated by sourcing/dossier_generator.py -- a
        # per-BRIEF procurement-fit assessment (sourcing/sourcing_agent.py),
        # deliberately separate from every other field in this bundle
        # (which answer "what evidence exists," not "does this supplier
        # fit what THIS buyer asked for"). None on every field for a
        # supplier that was never run through a sourcing brief.
        "sourcing_dossier": {
            "oem_odm_notes": supplier.get("sourcing_oem_odm_notes"),
            "factory_notes": supplier.get("sourcing_factory_notes"),
            "engineering_notes": supplier.get("sourcing_engineering_notes"),
            "export_notes": supplier.get("sourcing_export_notes"),
            "volume_suitability": supplier.get("sourcing_volume_suitability"),
            "payment_terms_notes": supplier.get("sourcing_payment_terms_notes"),
            "verification_status": supplier.get("sourcing_verification_status"),
        },
    }
