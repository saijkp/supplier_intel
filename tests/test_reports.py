"""
tests/test_reports.py

Tests for reports/generator.py — Markdown report rendering and CSV
export.
"""

from __future__ import annotations

import csv

from reports.generator import (
    CSV_COLUMNS,
    EXCEL_COLUMNS,
    SOURCING_CSV_COLUMNS,
    export_suppliers_csv,
    export_suppliers_excel,
    generate_markdown_report,
    save_markdown_report,
    suppliers_to_csv_string,
    suppliers_to_excel_bytes,
    suppliers_to_sourcing_csv_string,
)
from storage.database import initialise_schema
from storage.repository import SupplierRepository


def _make_repo(tmp_path):
    db_path = tmp_path / "test.db"
    initialise_schema(db_path)
    return SupplierRepository(db_path=db_path)


class TestGenerateMarkdownReport:

    def test_empty_database_reports_zero_suppliers(self, tmp_path):
        repo = _make_repo(tmp_path)
        content = generate_markdown_report(repo)
        assert "Suppliers: 0" in content
        assert "No suppliers matched" in content

    def test_report_includes_supplier_details(self, tmp_path):
        repo = _make_repo(tmp_path)
        supplier_id = repo.create_golden_record({
            "canonical_name": "Shenzhen LED Masters Co Ltd",
            "country": "China", "city": "Shenzhen",
            "is_manufacturer": True, "uscc_verified": True,
            "iso_9001": True, "e_mark_certified": True,
            "confirmed_shipments_uk": 12,
            "contact_name": "Li Wei", "primary_email": "sales@ledmasters.com",
        })
        repo.update_scores(supplier_id, {
            "composite_score": 82, "recommendation": "recommended",
            "verification_score": 90, "export_score": 80, "platform_score": 70, "contact_score": 100,
        })

        content = generate_markdown_report(repo)

        assert "Shenzhen LED Masters Co Ltd" in content
        assert "recommended" in content
        assert "82/100" in content
        assert "China, Shenzhen" in content
        assert "Manufacturer:** Yes" in content
        assert "ISO 9001" in content
        assert "E-mark" in content
        assert "Li Wei" in content
        assert "sales@ledmasters.com" in content

    def test_report_respects_recommendation_filter(self, tmp_path):
        repo = _make_repo(tmp_path)
        id_a = repo.create_golden_record({"canonical_name": "A Co"})
        id_b = repo.create_golden_record({"canonical_name": "B Co"})
        repo.update_scores(id_a, {"composite_score": 90, "recommendation": "recommended"})
        repo.update_scores(id_b, {"composite_score": 10, "recommendation": "avoid"})

        content = generate_markdown_report(repo, recommendation="recommended")
        assert "A Co" in content
        assert "B Co" not in content

    def test_report_respects_min_score_filter(self, tmp_path):
        repo = _make_repo(tmp_path)
        id_a = repo.create_golden_record({"canonical_name": "High Co"})
        id_b = repo.create_golden_record({"canonical_name": "Low Co"})
        repo.update_scores(id_a, {"composite_score": 90})
        repo.update_scores(id_b, {"composite_score": 5})

        content = generate_markdown_report(repo, min_score=50)
        assert "High Co" in content
        assert "Low Co" not in content

    def test_report_shows_query_when_given(self, tmp_path):
        repo = _make_repo(tmp_path)
        content = generate_markdown_report(repo, query="LED marker light")
        assert 'Search: "LED marker light"' in content

    def test_unknown_manufacturer_status_rendered(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({"canonical_name": "Unknown Status Co"})
        content = generate_markdown_report(repo)
        assert "Manufacturer:** Unknown" in content

    def test_confirmed_trader_rendered_as_no(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({"canonical_name": "Trader Co", "is_manufacturer": False})
        content = generate_markdown_report(repo)
        assert "Manufacturer:** No" in content

    def test_manufacturer_confidence_and_signals_rendered(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({
            "canonical_name": "Foo Co", "is_manufacturer": True,
            "manufacturer_confidence": 85,
            "manufacturer_signals": ["Registered business scope explicitly includes production/manufacturing"],
        })
        content = generate_markdown_report(repo)
        assert "confidence: 85/100" in content
        assert "Manufacturer evidence:" in content
        assert "explicitly includes production/manufacturing" in content


class TestSaveMarkdownReport:

    def test_save_writes_file(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({"canonical_name": "Foo Co"})

        output_path = tmp_path / "reports" / "suppliers.md"
        result_path = save_markdown_report(repo, output_path)

        assert result_path == output_path
        assert output_path.exists()
        assert "Foo Co" in output_path.read_text(encoding="utf-8")


class TestCSVExport:

    def test_suppliers_to_csv_string_includes_header_and_rows(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({
            "canonical_name": "Foo Co", "country": "China", "is_manufacturer": True,
        })
        suppliers = repo.list_suppliers()

        csv_text = suppliers_to_csv_string(suppliers)
        reader = csv.DictReader(csv_text.splitlines())
        rows = list(reader)

        assert reader.fieldnames == CSV_COLUMNS
        assert len(rows) == 1
        assert rows[0]["canonical_name"] == "Foo Co"
        assert rows[0]["country"] == "China"

    def test_export_suppliers_csv_writes_file(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({"canonical_name": "Foo Co", "country": "China"})
        repo.create_golden_record({"canonical_name": "Bar Co", "country": "India"})

        output_path = tmp_path / "exports" / "suppliers.csv"
        result_path = export_suppliers_csv(repo, output_path)

        assert result_path.exists()
        with open(result_path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2
        names = {row["canonical_name"] for row in rows}
        assert names == {"Foo Co", "Bar Co"}

    def test_export_respects_filters(self, tmp_path):
        repo = _make_repo(tmp_path)
        id_a = repo.create_golden_record({"canonical_name": "High Co"})
        id_b = repo.create_golden_record({"canonical_name": "Low Co"})
        repo.update_scores(id_a, {"composite_score": 90, "recommendation": "recommended"})
        repo.update_scores(id_b, {"composite_score": 5, "recommendation": "avoid"})

        output_path = tmp_path / "filtered.csv"
        export_suppliers_csv(repo, output_path, recommendation="recommended")

        with open(output_path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 1
        assert rows[0]["canonical_name"] == "High Co"

    def test_csv_columns_are_a_curated_subset(self):
        # Sanity check that this stays a deliberately curated list, not
        # every column in the suppliers table. Threshold bumped from 30
        # to 35 when ai_confidence_score/procurement_recommendation
        # were added (Procurement Decision Engine foundation) --
        # legitimate growth, not scope creep back toward "every column."
        assert "canonical_name" in CSV_COLUMNS
        assert "id" in CSV_COLUMNS
        assert len(CSV_COLUMNS) < 35

    def test_manufacturer_signals_flattened_to_readable_string(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({
            "canonical_name": "Foo Co",
            "manufacturer_confidence": 90,
            "manufacturer_signals": ["Signal one", "Signal two"],
        })
        csv_text = suppliers_to_csv_string(repo.list_suppliers())
        reader = csv.DictReader(csv_text.splitlines())
        row = next(reader)

        assert row["manufacturer_confidence"] == "90"
        assert "Signal one" in row["manufacturer_signals"]
        assert "Signal two" in row["manufacturer_signals"]
        assert "[" not in row["manufacturer_signals"]  # not a raw Python list repr

    def test_address_is_included(self, tmp_path):
        """address was missing from CSV_COLUMNS despite being one of
        the fields buyers most need -- regression guard against
        losing it again."""
        assert "address" in CSV_COLUMNS


class TestSourcingCSVExport:

    def test_key_contacts_split_into_three_derived_columns(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({
            "canonical_name": "Foo Co", "domain": "foo.example.com",
            "key_contacts": [
                {"name": "Jane Doe", "title": "Procurement Manager", "email": "jane@foo.example.com",
                 "phone": None, "linkedin_url": None, "role_category": "procurement"},
                {"name": "John Smith", "title": "CEO", "email": "john@foo.example.com",
                 "phone": None, "linkedin_url": None, "role_category": "ceo"},
            ],
        })
        suppliers = repo.list_suppliers()

        csv_text = suppliers_to_sourcing_csv_string(suppliers)
        reader = csv.DictReader(csv_text.splitlines())
        row = next(reader)

        assert reader.fieldnames == SOURCING_CSV_COLUMNS
        assert "Jane Doe" in row["procurement_manager"]
        assert "jane@foo.example.com" in row["procurement_manager"]
        assert "John Smith" in row["ceo"]
        assert row["sales_manager"] == ""  # no matching role -- left blank, never fabricated

    def test_supplier_with_no_contacts_leaves_columns_blank(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({"canonical_name": "No Contacts Co"})
        suppliers = repo.list_suppliers()

        csv_text = suppliers_to_sourcing_csv_string(suppliers)
        row = next(csv.DictReader(csv_text.splitlines()))

        assert row["procurement_manager"] == ""
        assert row["sales_manager"] == ""
        assert row["ceo"] == ""

    def test_only_first_contact_per_role_is_used(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({
            "canonical_name": "Foo Co",
            "key_contacts": [
                {"name": "First Buyer", "title": "Procurement Manager", "email": None,
                 "phone": None, "linkedin_url": None, "role_category": "procurement"},
                {"name": "Second Buyer", "title": "Purchasing Director", "email": None,
                 "phone": None, "linkedin_url": None, "role_category": "procurement"},
            ],
        })
        suppliers = repo.list_suppliers()

        csv_text = suppliers_to_sourcing_csv_string(suppliers)
        row = next(csv.DictReader(csv_text.splitlines()))

        assert "First Buyer" in row["procurement_manager"]
        assert "Second Buyer" not in row["procurement_manager"]


class TestExcelExport:

    def _read_rows(self, excel_bytes: bytes):
        from io import BytesIO

        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(excel_bytes))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header, data_rows = rows[0], rows[1:]
        return header, [dict(zip(header, row)) for row in data_rows]

    def test_suppliers_to_excel_bytes_includes_header_and_rows(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({
            "canonical_name": "Foo Co", "country": "China", "is_manufacturer": True,
        })
        excel_bytes = suppliers_to_excel_bytes(repo.list_suppliers())

        header, rows = self._read_rows(excel_bytes)
        assert list(header) == EXCEL_COLUMNS
        assert len(rows) == 1
        assert rows[0]["canonical_name"] == "Foo Co"
        assert rows[0]["country"] == "China"

    def test_excel_columns_are_a_superset_of_csv_columns(self):
        assert set(CSV_COLUMNS).issubset(set(EXCEL_COLUMNS))
        assert "secondary_emails" in EXCEL_COLUMNS
        assert "contact_form_url" in EXCEL_COLUMNS
        assert "facility_address_verified" in EXCEL_COLUMNS
        assert "linkedin_url" in EXCEL_COLUMNS

    def test_secondary_emails_flattened_to_readable_string(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({
            "canonical_name": "Foo Co",
            "primary_email": "sales@foo.com",
            "secondary_emails": ["info@foo.com", "support@foo.com"],
        })
        excel_bytes = suppliers_to_excel_bytes(repo.list_suppliers())
        _, rows = self._read_rows(excel_bytes)

        assert "info@foo.com" in rows[0]["secondary_emails"]
        assert "support@foo.com" in rows[0]["secondary_emails"]
        assert "[" not in rows[0]["secondary_emails"]

    def test_export_suppliers_excel_writes_a_readable_file(self, tmp_path):
        repo = _make_repo(tmp_path)
        repo.create_golden_record({"canonical_name": "Foo Co", "country": "China"})
        repo.create_golden_record({"canonical_name": "Bar Co", "country": "India"})

        output_path = tmp_path / "exports" / "suppliers.xlsx"
        result_path = export_suppliers_excel(repo, output_path)

        assert result_path.exists()
        header, rows = self._read_rows(result_path.read_bytes())
        assert len(rows) == 2
        names = {row["canonical_name"] for row in rows}
        assert names == {"Foo Co", "Bar Co"}

    def test_export_respects_filters(self, tmp_path):
        repo = _make_repo(tmp_path)
        id_a = repo.create_golden_record({"canonical_name": "High Co"})
        id_b = repo.create_golden_record({"canonical_name": "Low Co"})
        repo.update_scores(id_a, {"composite_score": 90, "recommendation": "recommended"})
        repo.update_scores(id_b, {"composite_score": 5, "recommendation": "avoid"})

        output_path = tmp_path / "filtered.xlsx"
        export_suppliers_excel(repo, output_path, recommendation="recommended")

        _, rows = self._read_rows(output_path.read_bytes())
        assert len(rows) == 1
        assert rows[0]["canonical_name"] == "High Co"

    def test_empty_supplier_list_produces_header_only(self, tmp_path):
        repo = _make_repo(tmp_path)
        excel_bytes = suppliers_to_excel_bytes(repo.list_suppliers())
        header, rows = self._read_rows(excel_bytes)
        assert list(header) == EXCEL_COLUMNS
        assert rows == []
