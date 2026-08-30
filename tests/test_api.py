"""
tests/test_api.py

Tests for api/app.py, using FastAPI's TestClient (real HTTP requests
in-process, no network) against a real, temporary SQLite database --
not mocks of the repository. This exercises the actual dependency
injection, auth, and JSON serialisation path a real client would hit.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

TOKEN = "test-token-123"


@pytest.fixture()
def client(tmp_path, monkeypatch):
    """Two fixes over a naive TestClient setup, both for real reasons
    found while first writing these tests:

    1. `config.settings.DB_PATH` (and CORS's `ALLOWED_ORIGINS`) are
       resolved once at module-import time from the environment --
       monkeypatching the env var after `api.app` has already been
       imported does nothing. FastAPI's own `app.dependency_overrides`
       is the correct, import-order-independent way to inject a test
       database, so that's what this uses instead of env manipulation.

    2. FastAPI's TestClient actually executes `BackgroundTasks` for
       real, synchronously, within the `with TestClient(...)` block --
       not a no-op or a mock. Left unpatched, every test that creates a
       pipeline job would trigger a real, live pipeline run: real
       network calls to Alibaba/HKTDC/Volza/etc., ~80 seconds per test,
       and — if a real paid key were ever present in a test
       environment — real spend. `api.app.run_pipeline_job` (the name
       bound in api.app's own namespace via `from api.jobs import
       run_pipeline_job`, not `api.jobs.run_pipeline_job` itself, which
       a caller who already imported it would not see change) is
       patched to a fast, no-network fake for every test that only
       needs to prove the API triggers a job, not that the underlying
       pipeline works -- that's already covered by
       tests/test_capability_pipeline_stage.py and friends.
    """
    db_path = tmp_path / "test_api.db"

    import api.auth
    monkeypatch.setattr(api.auth, "API_ACCESS_TOKEN", TOKEN)

    import api.app
    from storage.database import initialise_schema
    from storage.repository import SupplierRepository

    initialise_schema(db_path)
    test_repo = SupplierRepository(db_path=db_path)
    api.app.app.dependency_overrides[api.app.get_repo] = lambda: test_repo

    def fake_run_pipeline_job(job_id, query, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"scraped": 0})

    monkeypatch.setattr(api.app, "run_pipeline_job", fake_run_pipeline_job)

    def fake_run_enrichment_job(job_id, stage, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"stage": stage})

    monkeypatch.setattr(api.app, "run_enrichment_job", fake_run_enrichment_job)

    def fake_run_collection_job(job_id, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"attempted": 0})

    monkeypatch.setattr(api.app, "run_collection_job", fake_run_collection_job)

    def fake_run_verification_job(job_id, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"confidence_score": 0})

    monkeypatch.setattr(api.app, "run_verification_job", fake_run_verification_job)

    def fake_run_reverify_job(job_id, supplier_id):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"supplier_id": supplier_id})

    monkeypatch.setattr(api.app, "run_reverify_job", fake_run_reverify_job)

    def fake_run_discovery_job(job_id, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"candidates_found": 0})

    monkeypatch.setattr(api.app, "run_discovery_job", fake_run_discovery_job)

    def fake_run_sourcing_job(job_id, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"run_id": 0, "qualified_supplier_ids": []})

    monkeypatch.setattr(api.app, "run_sourcing_job", fake_run_sourcing_job)

    def fake_run_single_company_job(job_id, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"resolved_supplier_id": None})

    monkeypatch.setattr(api.app, "run_single_company_job", fake_run_single_company_job)

    def fake_run_supplier_correction_job(job_id, supplier_id, options):
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"supplier_id": supplier_id, "status": "needs_url"})

    monkeypatch.setattr(api.app, "run_supplier_correction_job", fake_run_supplier_correction_job)

    batch_job_calls = []

    def fake_run_batch_job(job_id, csv_bytes, recover_dead_domains=False, recovery_product_term=None,
                            default_region=None):
        batch_job_calls.append({
            "job_id": job_id, "recover_dead_domains": recover_dead_domains,
            "recovery_product_term": recovery_product_term, "default_region": default_region,
        })
        test_repo.mark_pipeline_job_running(job_id)
        test_repo.mark_pipeline_job_completed(job_id, stats={"processed": 0})

    monkeypatch.setattr(api.app, "run_batch_job", fake_run_batch_job)

    with TestClient(api.app.app) as test_client:
        test_client.repo = test_repo
        test_client.batch_job_calls = batch_job_calls
        yield test_client

    api.app.app.dependency_overrides.clear()


def auth_headers():
    return {"Authorization": f"Bearer {TOKEN}"}


class TestHealth:

    def test_health_needs_no_auth(self, client):
        response = client.get("/health")
        assert response.status_code == 200
        assert response.json() == {"status": "ok"}


class TestAuth:

    def test_missing_token_is_rejected(self, client):
        response = client.get("/suppliers/search")
        assert response.status_code == 401

    def test_wrong_token_is_rejected(self, client):
        response = client.get("/suppliers/search", headers={"Authorization": "Bearer wrong"})
        assert response.status_code == 401

    def test_correct_token_is_accepted(self, client):
        response = client.get("/suppliers/search", headers=auth_headers())
        assert response.status_code == 200

    def test_no_token_configured_fails_closed(self, client, monkeypatch):
        """The specific security property that matters: an
        unconfigured server refuses requests rather than silently
        allowing everyone through."""
        import api.auth

        monkeypatch.setattr(api.auth, "API_ACCESS_TOKEN", None)
        response = client.get("/suppliers/search", headers=auth_headers())
        assert response.status_code == 503


class TestSearchEndpoint:

    def test_search_returns_a_seeded_supplier(self, client):
        client.repo.create_golden_record({
            "canonical_name": "Acme Trailer Parts", "country": "United Kingdom",
            "domain": "acme.example.com", "product_keywords": ["wheel hub"],
        })
        response = client.get(
            "/suppliers/search", params={"product": "wheel hub"}, headers=auth_headers(),
        )
        assert response.status_code == 200
        results = response.json()
        assert len(results) == 1
        assert results[0]["canonical_name"] == "Acme Trailer Parts"

    def test_search_with_no_matches_returns_empty_list_not_404(self, client):
        response = client.get(
            "/suppliers/search", params={"product": "nonexistent widget"}, headers=auth_headers(),
        )
        assert response.status_code == 200
        assert response.json() == []

    def test_export_and_capacity_fields_pass_through(self, client):
        """Compare/rank UI feed -- these DB columns existed long before
        this field set was wired into SupplierSearchResult; regression
        guard against them silently dropping out of the response
        again."""
        client.repo.create_golden_record({
            "canonical_name": "Acme Trailer Parts", "domain": "acme.example.com",
            "confirmed_shipments_uk": 3, "confirmed_shipments_eu": 1,
            "exports_to_uk": True, "active_export_countries": ["United Kingdom", "France"],
            "employee_count": "50-100", "factory_size_sqm": 5000,
        })
        response = client.get("/suppliers/search", headers=auth_headers())
        assert response.status_code == 200
        result = response.json()[0]
        assert result["confirmed_shipments_uk"] == 3
        assert result["confirmed_shipments_eu"] == 1
        assert result["exports_to_uk"] is True
        assert result["active_export_countries"] == ["United Kingdom", "France"]
        assert result["employee_count"] == "50-100"
        assert result["factory_size_sqm"] == 5000

    def test_include_capabilities_false_by_default_leaves_matched_capabilities_empty(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme", "domain": "acme.example.com"})
        client.repo.add_capability_finding(supplier_id, {
            "reported_term": "PPAP", "canonical_term": "ppap capability", "category": "oem_readiness",
            "relationship": "asserted", "confidence": 0.9, "evidence": "We support PPAP.", "source_url": "https://acme.example.com",
        })
        response = client.get("/suppliers/search", headers=auth_headers())
        assert response.status_code == 200
        assert response.json()[0]["matched_capabilities"] == []

    def test_include_capabilities_true_populates_matched_capabilities(self, client):
        """The Compare/rank UI's whole reason for existing -- OEM/
        engineering/certification evidence must be readable from a
        plain browse query, not only when filtering by `require`."""
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme", "domain": "acme.example.com"})
        client.repo.add_capability_finding(supplier_id, {
            "reported_term": "PPAP", "canonical_term": "ppap capability", "category": "oem_readiness",
            "relationship": "asserted", "confidence": 0.9, "evidence": "We support PPAP.", "source_url": "https://acme.example.com",
        })
        response = client.get(
            "/suppliers/search", params={"include_capabilities": "true"}, headers=auth_headers(),
        )
        assert response.status_code == 200
        capabilities = response.json()[0]["matched_capabilities"]
        assert len(capabilities) == 1
        assert capabilities[0]["canonical_term"] == "ppap capability"
        assert capabilities[0]["category"] == "oem_readiness"

    def test_country_filter_is_passed_through(self, client):
        client.repo.create_golden_record({
            "canonical_name": "UK Co", "country": "United Kingdom", "domain": "uk.example.com",
        })
        client.repo.create_golden_record({
            "canonical_name": "China Co", "country": "China", "domain": "cn.example.com",
        })
        response = client.get(
            "/suppliers/search", params={"country": "United Kingdom"}, headers=auth_headers(),
        )
        names = {r["canonical_name"] for r in response.json()}
        assert names == {"UK Co"}

    def test_repeated_require_params_are_all_applied(self, client):
        """Confirms query-string list handling (?require=a&require=b)
        works through FastAPI's own parsing, not just at the
        repository layer (already covered elsewhere)."""
        response = client.get(
            "/suppliers/search",
            params=[("require", "iso 9001"), ("require", "sub-assembly")],
            headers=auth_headers(),
        )
        assert response.status_code == 200

    def test_unrecognised_capability_returns_400_not_500(self, client):
        response = client.get(
            "/suppliers/search", params={"require": "not-a-real-capability"}, headers=auth_headers(),
        )
        assert response.status_code == 400
        assert "not a recognised capability" in response.json()["detail"]

    def test_is_manufacturer_boolean_is_correctly_coerced_from_sqlite_int(self, client):
        """SQLite stores booleans as 0/1 -- this proves the API
        response actually comes back as a real JSON boolean, not 1."""
        supplier_id = client.repo.create_golden_record({
            "canonical_name": "Verified Co", "domain": "verified.example.com",
        })
        client.repo.update_supplier_fields(supplier_id, {"is_manufacturer": True})
        response = client.get(
            "/suppliers/search", params={"product": "Verified"}, headers=auth_headers(),
        )
        results = response.json()
        assert len(results) == 1
        assert results[0]["is_manufacturer"] is True


class TestGetSupplierEndpoint:

    def test_returns_full_detail_including_capabilities(self, client):
        supplier_id = client.repo.create_golden_record({
            "canonical_name": "Acme", "domain": "acme.example.com",
        })
        client.repo.add_capability_finding(supplier_id, {
            "reported_term": "iso 9001", "canonical_term": "iso 9001", "category": "standard",
            "relationship": "in_house", "confidence": 0.9,
            "evidence": "we are ISO 9001 certified", "source_url": "https://acme.example.com",
        })
        response = client.get(f"/suppliers/{supplier_id}", headers=auth_headers())
        assert response.status_code == 200
        body = response.json()
        assert body["canonical_name"] == "Acme"
        assert len(body["matched_capabilities"]) == 1
        assert body["matched_capabilities"][0]["canonical_term"] == "iso 9001"

    def test_nonexistent_supplier_returns_404(self, client):
        response = client.get("/suppliers/999999", headers=auth_headers())
        assert response.status_code == 404


class TestPipelineJobEndpoints:

    def test_creating_a_job_returns_202_with_a_job_id(self, client):
        response = client.post(
            "/pipeline/jobs", json={"query": "wheel bearings"}, headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "wheel bearings"
        assert body["id"]

    def test_created_job_is_retrievable_by_id(self, client):
        create_response = client.post(
            "/pipeline/jobs", json={"query": "led lighting"}, headers=auth_headers(),
        )
        job_id = create_response.json()["id"]
        get_response = client.get(f"/pipeline/jobs/{job_id}", headers=auth_headers())
        assert get_response.status_code == 200
        assert get_response.json()["query"] == "led lighting"

    def test_nonexistent_job_returns_404(self, client):
        response = client.get("/pipeline/jobs/does-not-exist", headers=auth_headers())
        assert response.status_code == 404

    def test_list_jobs_returns_created_jobs(self, client):
        client.post("/pipeline/jobs", json={"query": "wheel bearings"}, headers=auth_headers())
        client.post("/pipeline/jobs", json={"query": "led lighting"}, headers=auth_headers())
        response = client.get("/pipeline/jobs", headers=auth_headers())
        assert response.status_code == 200
        assert len(response.json()) == 2

    def test_missing_query_is_a_validation_error(self, client):
        response = client.post("/pipeline/jobs", json={}, headers=auth_headers())
        assert response.status_code == 422


class TestEnrichmentJobEndpoints:

    def test_creating_an_enrichment_job_returns_202_with_a_job_id(self, client):
        response = client.post(
            "/pipeline/enrichment-jobs", json={"stage": "find_websites"}, headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[enrichment] find_websites"
        assert body["id"]

    def test_created_enrichment_job_is_retrievable_by_id(self, client):
        create_response = client.post(
            "/pipeline/enrichment-jobs", json={"stage": "extract_capabilities", "limit": 10},
            headers=auth_headers(),
        )
        job_id = create_response.json()["id"]
        get_response = client.get(f"/pipeline/jobs/{job_id}", headers=auth_headers())
        assert get_response.status_code == 200
        assert get_response.json()["query"] == "[enrichment] extract_capabilities"

    def test_missing_stage_is_a_validation_error(self, client):
        response = client.post("/pipeline/enrichment-jobs", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_defaults_match_documented_cost_safe_behaviour(self, client):
        response = client.post(
            "/pipeline/enrichment-jobs", json={"stage": "verify_facilities"}, headers=auth_headers(),
        )
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["force"] is False
        assert job["options"]["limit"] is None
        assert job["options"]["assess_photos"] is False


class TestCollectionJobEndpoints:

    def test_creating_a_supplier_id_job_returns_202_with_a_job_id(self, client):
        response = client.post(
            "/collection/jobs", json={"supplier_id": 5}, headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[collection] supplier #5"
        assert body["id"]

    def test_creating_a_pending_batch_job(self, client):
        response = client.post(
            "/collection/jobs", json={"pending": True, "limit": 10}, headers=auth_headers(),
        )
        assert response.status_code == 202
        assert response.json()["query"] == "[collection] pending batch"

    def test_created_job_is_retrievable_by_id(self, client):
        create_response = client.post("/collection/jobs", json={"supplier_id": 5}, headers=auth_headers())
        job_id = create_response.json()["id"]
        get_response = client.get(f"/pipeline/jobs/{job_id}", headers=auth_headers())
        assert get_response.status_code == 200
        assert get_response.json()["query"] == "[collection] supplier #5"

    def test_neither_supplier_id_nor_pending_is_a_validation_error(self, client):
        response = client.post("/collection/jobs", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_requires_auth(self, client):
        response = client.post("/collection/jobs", json={"supplier_id": 5})
        assert response.status_code == 401


class TestBatchUploadEndpoint:

    _CSV = b"company_name,website\nAcme Co,https://acme.example.com\n"

    def test_plain_upload_still_works(self, client):
        response = client.post(
            "/batch/upload", files={"file": ("suppliers.csv", self._CSV, "text/csv")}, headers=auth_headers(),
        )
        assert response.status_code == 202
        assert client.batch_job_calls[-1]["recover_dead_domains"] is False
        assert client.batch_job_calls[-1]["recovery_product_term"] is None
        assert client.batch_job_calls[-1]["default_region"] is None

    def test_empty_file_is_rejected(self, client):
        response = client.post(
            "/batch/upload", files={"file": ("suppliers.csv", b"", "text/csv")}, headers=auth_headers(),
        )
        assert response.status_code == 422

    def test_requires_auth(self, client):
        response = client.post("/batch/upload", files={"file": ("suppliers.csv", self._CSV, "text/csv")})
        assert response.status_code == 401

    def test_recover_dead_domains_without_product_term_fails_closed(self, client):
        response = client.post(
            "/batch/upload", files={"file": ("suppliers.csv", self._CSV, "text/csv")},
            data={"recover_dead_domains": "true"}, headers=auth_headers(),
        )
        assert response.status_code == 422
        assert client.batch_job_calls == []

    def test_recover_dead_domains_with_product_term_is_accepted_and_threaded_through(self, client):
        response = client.post(
            "/batch/upload", files={"file": ("suppliers.csv", self._CSV, "text/csv")},
            data={"recover_dead_domains": "true", "recovery_product_term": "trailer axle"},
            headers=auth_headers(),
        )
        assert response.status_code == 202
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["recover_dead_domains"] is True
        assert job["options"]["recovery_product_term"] == "trailer axle"
        call = client.batch_job_calls[-1]
        assert call["recover_dead_domains"] is True
        assert call["recovery_product_term"] == "trailer axle"

    def test_default_region_is_accepted_and_threaded_through(self, client):
        response = client.post(
            "/batch/upload", files={"file": ("suppliers.csv", self._CSV, "text/csv")},
            data={"default_region": "GB"}, headers=auth_headers(),
        )
        assert response.status_code == 202
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["default_region"] == "GB"
        assert client.batch_job_calls[-1]["default_region"] == "GB"

    def test_default_region_is_independent_of_recover_dead_domains(self, client):
        """Not gated behind recover_dead_domains -- a plain upload with
        only a region hint set must not be rejected the way an
        unpaired recover_dead_domains would be."""
        response = client.post(
            "/batch/upload", files={"file": ("suppliers.csv", self._CSV, "text/csv")},
            data={"default_region": "GB"}, headers=auth_headers(),
        )
        assert response.status_code == 202
        assert client.batch_job_calls[-1]["recover_dead_domains"] is False


class TestContactsJobEndpoints:

    def test_creating_a_supplier_id_job_returns_202_with_a_job_id(self, client):
        response = client.post(
            "/contacts/jobs", json={"supplier_id": 5}, headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[contacts] supplier #5"
        assert body["id"]

    def test_creating_a_pending_batch_job(self, client):
        response = client.post(
            "/contacts/jobs", json={"pending": True, "limit": 10}, headers=auth_headers(),
        )
        assert response.status_code == 202
        assert response.json()["query"] == "[contacts] pending batch"

    def test_created_job_is_retrievable_by_id(self, client):
        create_response = client.post("/contacts/jobs", json={"supplier_id": 5}, headers=auth_headers())
        job_id = create_response.json()["id"]
        get_response = client.get(f"/pipeline/jobs/{job_id}", headers=auth_headers())
        assert get_response.status_code == 200
        assert get_response.json()["query"] == "[contacts] supplier #5"

    def test_neither_supplier_id_nor_pending_is_a_validation_error(self, client):
        response = client.post("/contacts/jobs", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_requires_auth(self, client):
        response = client.post("/contacts/jobs", json={"supplier_id": 5})
        assert response.status_code == 401


class TestFactoryFactsJobEndpoints:

    def test_creating_a_supplier_id_job_returns_202_with_a_job_id(self, client):
        response = client.post(
            "/factory-facts/jobs", json={"supplier_id": 5}, headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[factory-facts] supplier #5"
        assert body["id"]

    def test_creating_a_pending_batch_job(self, client):
        response = client.post(
            "/factory-facts/jobs", json={"pending": True, "limit": 10}, headers=auth_headers(),
        )
        assert response.status_code == 202
        assert response.json()["query"] == "[factory-facts] pending batch"

    def test_created_job_is_retrievable_by_id(self, client):
        create_response = client.post("/factory-facts/jobs", json={"supplier_id": 5}, headers=auth_headers())
        job_id = create_response.json()["id"]
        get_response = client.get(f"/pipeline/jobs/{job_id}", headers=auth_headers())
        assert get_response.status_code == 200
        assert get_response.json()["query"] == "[factory-facts] supplier #5"

    def test_neither_supplier_id_nor_pending_is_a_validation_error(self, client):
        response = client.post("/factory-facts/jobs", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_requires_auth(self, client):
        response = client.post("/factory-facts/jobs", json={"supplier_id": 5})
        assert response.status_code == 401


class TestVerificationJobEndpoints:

    def test_creating_a_supplier_id_job_returns_202_with_a_job_id(self, client):
        response = client.post("/verification/jobs", json={"supplier_id": 5}, headers=auth_headers())
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[verification] supplier #5"

    def test_creating_a_pending_batch_job(self, client):
        response = client.post("/verification/jobs", json={"pending": True}, headers=auth_headers())
        assert response.status_code == 202
        assert response.json()["query"] == "[verification] pending batch"

    def test_neither_supplier_id_nor_pending_is_a_validation_error(self, client):
        response = client.post("/verification/jobs", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_requires_auth(self, client):
        response = client.post("/verification/jobs", json={"supplier_id": 5})
        assert response.status_code == 401


class TestDiscoveryJobEndpoints:

    def test_creating_a_job_returns_202_with_a_job_id(self, client):
        response = client.post("/discovery/jobs", json={"product": "trailer axle"}, headers=auth_headers())
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[discovery] trailer axle"
        assert body["id"]

    def test_created_job_is_retrievable_by_id(self, client):
        create_response = client.post(
            "/discovery/jobs", json={"product": "trailer axle", "country": "China", "max_candidates": 10},
            headers=auth_headers(),
        )
        job_id = create_response.json()["id"]
        get_response = client.get(f"/pipeline/jobs/{job_id}", headers=auth_headers())
        assert get_response.status_code == 200
        assert get_response.json()["query"] == "[discovery] trailer axle"

    def test_missing_product_is_a_validation_error(self, client):
        response = client.post("/discovery/jobs", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_defaults_match_documented_behaviour(self, client):
        response = client.post("/discovery/jobs", json={"product": "trailer axle"}, headers=auth_headers())
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["category"] is None
        assert job["options"]["country"] is None
        assert job["options"]["max_candidates"] == 20

    def test_requires_auth(self, client):
        response = client.post("/discovery/jobs", json={"product": "trailer axle"})
        assert response.status_code == 401

    def test_target_count_and_max_multiplier_accepted_and_stored(self, client):
        response = client.post(
            "/discovery/jobs", json={"product": "forklift", "target_count": 15, "max_multiplier": 4},
            headers=auth_headers(),
        )
        assert response.status_code == 202
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["target_count"] == 15
        assert job["options"]["max_multiplier"] == 4

    def test_target_count_defaults_to_none_and_max_multiplier_to_five(self, client):
        response = client.post("/discovery/jobs", json={"product": "trailer axle"}, headers=auth_headers())
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["target_count"] is None
        assert job["options"]["max_multiplier"] == 5


class TestSingleCompanyEnrichEndpoint:

    def test_creating_a_job_returns_202_with_a_job_id(self, client):
        response = client.post(
            "/companies/enrich", json={"input_text": "Acme Trailer Co"}, headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[single-company] Acme Trailer Co"
        assert body["id"]

    def test_created_job_is_retrievable_by_id(self, client):
        create_response = client.post(
            "/companies/enrich", json={"input_text": "acmetrailer.com"}, headers=auth_headers(),
        )
        job_id = create_response.json()["id"]
        get_response = client.get(f"/pipeline/jobs/{job_id}", headers=auth_headers())
        assert get_response.status_code == 200
        assert get_response.json()["query"] == "[single-company] acmetrailer.com"

    def test_missing_input_text_is_a_validation_error(self, client):
        response = client.post("/companies/enrich", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_country_is_optional(self, client):
        response = client.post("/companies/enrich", json={"input_text": "Acme Trailer Co"}, headers=auth_headers())
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["country"] is None

    def test_requires_auth(self, client):
        response = client.post("/companies/enrich", json={"input_text": "Acme Trailer Co"})
        assert response.status_code == 401


class TestSupplierCorrectionEndpoint:

    def test_creating_a_job_returns_202_with_a_job_id(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Ashpock", "domain": "shpock.com"})
        response = client.post(
            f"/suppliers/{supplier_id}/correct-domain",
            json={"reason": "false match: shpock.com is an unrelated app"},
            headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == f"[correct-supplier] #{supplier_id}"
        assert body["id"]

    def test_reason_is_optional(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Ashpock", "domain": "shpock.com"})
        response = client.post(f"/suppliers/{supplier_id}/correct-domain", json={}, headers=auth_headers())
        assert response.status_code == 202

    def test_missing_supplier_is_404(self, client):
        response = client.post("/suppliers/999999/correct-domain", json={}, headers=auth_headers())
        assert response.status_code == 404

    def test_requires_auth(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Ashpock", "domain": "shpock.com"})
        response = client.post(f"/suppliers/{supplier_id}/correct-domain", json={})
        assert response.status_code == 401

    def test_domain_and_canonical_name_pass_through_to_job_options(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Ashpock", "domain": "shpock.com"})
        response = client.post(
            f"/suppliers/{supplier_id}/correct-domain",
            json={"domain": "aspoeck.com", "canonical_name": "Aspoeck Systems"},
            headers=auth_headers(),
        )
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["domain"] == "aspoeck.com"
        assert job["options"]["canonical_name"] == "Aspoeck Systems"

    def test_flag_reason_passes_through_to_job_options(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Ashpock"})
        response = client.post(
            f"/suppliers/{supplier_id}/correct-domain",
            json={"flag_reason": "duplicate of #123 (Aspoeck Systems, aspoeck.com)"},
            headers=auth_headers(),
        )
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["flag_reason"] == "duplicate of #123 (Aspoeck Systems, aspoeck.com)"


class TestBackfillDiscoveryProductKeywordsEndpoint:

    def test_backfills_supplier_created_by_a_completed_discovery_job(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Winch Co"})
        client.repo.create_pipeline_job(job_id="job-1", query="[discovery] winch", options={})
        client.repo.mark_pipeline_job_completed(
            "job-1", stats={"new_supplier_ids": [supplier_id], "candidates_found": 1},
        )

        response = client.post("/discovery/backfill-product-keywords", headers=auth_headers())

        assert response.status_code == 200
        body = response.json()
        assert body["updated_count"] == 1
        assert body["updated_supplier_ids"] == [supplier_id]
        supplier = client.repo.get_supplier(supplier_id)
        assert supplier["product_keywords"] == ["winch"]

    def test_runs_synchronously_not_through_the_job_queue(self, client):
        """No BackgroundTasks, no pipeline_jobs row created by this
        call -- it's a fast, pure DB read/write, not a paid-API job."""
        response = client.post("/discovery/backfill-product-keywords", headers=auth_headers())
        assert response.status_code == 200
        assert "id" not in response.json()

    def test_requires_auth(self, client):
        response = client.post("/discovery/backfill-product-keywords")
        assert response.status_code == 401


class TestReverifyEndpoint:

    def test_creates_a_job_for_an_existing_supplier(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme", "domain": "acme.example.com"})
        response = client.post(f"/suppliers/{supplier_id}/reverify", headers=auth_headers())
        assert response.status_code == 202
        assert response.json()["query"] == f"[reverify] supplier #{supplier_id}"

    def test_unknown_supplier_returns_404(self, client):
        response = client.post("/suppliers/999999/reverify", headers=auth_headers())
        assert response.status_code == 404

    def test_requires_auth(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme"})
        response = client.post(f"/suppliers/{supplier_id}/reverify")
        assert response.status_code == 401


class TestCreateSourcingRunEndpoint:

    def test_creates_a_job_with_the_sourcing_label(self, client):
        response = client.post(
            "/sourcing/runs", json={"brief_text": "find 10 winch manufacturers in China"},
            headers=auth_headers(),
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["query"] == "[sourcing] find 10 winch manufacturers in China"
        assert body["id"]

    def test_long_brief_text_is_truncated_in_the_job_label(self, client):
        long_brief = "find genuine manufacturers " + "for a very specific niche product " * 5
        response = client.post("/sourcing/runs", json={"brief_text": long_brief}, headers=auth_headers())
        query = response.json()["query"]
        assert query.startswith("[sourcing] ")
        assert len(query) <= len("[sourcing] ") + 60

    def test_missing_brief_text_is_a_validation_error(self, client):
        response = client.post("/sourcing/runs", json={}, headers=auth_headers())
        assert response.status_code == 422

    def test_default_max_multiplier_is_recorded_in_job_options(self, client):
        response = client.post(
            "/sourcing/runs", json={"brief_text": "find winch manufacturers"}, headers=auth_headers(),
        )
        job_id = response.json()["id"]
        job = client.repo.get_pipeline_job(job_id)
        assert job["options"]["max_multiplier"] == 5

    def test_requires_auth(self, client):
        response = client.post("/sourcing/runs", json={"brief_text": "find winch manufacturers"})
        assert response.status_code == 401


class TestGetSourcingRunEndpoint:

    def test_returns_run_with_resolved_qualified_suppliers(self, client):
        id_a = client.repo.create_golden_record({"canonical_name": "Acme Winch Co", "country": "China"})
        id_b = client.repo.create_golden_record({"canonical_name": "Best Winch Co", "country": "India"})
        run_id = client.repo.record_sourcing_run(brief_text="find winch manufacturers", target_count=2)
        client.repo.complete_sourcing_run(run_id, qualified_supplier_ids=[id_a, id_b], examined_count=5)

        response = client.get(f"/sourcing/runs/{run_id}", headers=auth_headers())

        assert response.status_code == 200
        body = response.json()
        assert body["status"] == "completed"
        assert body["examined_count"] == 5
        assert {s["id"] for s in body["qualified_suppliers"]} == {id_a, id_b}

    def test_running_run_with_no_qualified_suppliers_yet_returns_empty_list(self, client):
        run_id = client.repo.record_sourcing_run(brief_text="find winch manufacturers", target_count=5)

        response = client.get(f"/sourcing/runs/{run_id}", headers=auth_headers())

        assert response.status_code == 200
        body = response.json()
        assert body["status"] == "running"
        assert body["qualified_suppliers"] == []

    def test_unknown_run_returns_404(self, client):
        response = client.get("/sourcing/runs/999999", headers=auth_headers())
        assert response.status_code == 404

    def test_requires_auth(self, client):
        run_id = client.repo.record_sourcing_run(brief_text="find winch manufacturers", target_count=5)
        response = client.get(f"/sourcing/runs/{run_id}")
        assert response.status_code == 401


class TestSourcingRunCsvExportEndpoint:

    def test_exports_only_the_qualified_suppliers(self, client):
        id_a = client.repo.create_golden_record({
            "canonical_name": "Acme Winch Co", "country": "China", "primary_email": "sales@acme.example.com",
        })
        client.repo.create_golden_record({"canonical_name": "Not Qualified Co"})
        run_id = client.repo.record_sourcing_run(brief_text="find winch manufacturers", target_count=1)
        client.repo.complete_sourcing_run(run_id, qualified_supplier_ids=[id_a], examined_count=3)

        response = client.get(f"/sourcing/runs/{run_id}/export.csv", headers=auth_headers())

        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "Acme Winch Co" in response.text
        assert "sales@acme.example.com" in response.text
        assert "Not Qualified Co" not in response.text

    def test_unknown_run_returns_404(self, client):
        response = client.get("/sourcing/runs/999999/export.csv", headers=auth_headers())
        assert response.status_code == 404

    def test_run_with_no_qualified_suppliers_exports_header_only(self, client):
        run_id = client.repo.record_sourcing_run(brief_text="find winch manufacturers", target_count=5)
        client.repo.complete_sourcing_run(run_id, qualified_supplier_ids=[], examined_count=5)

        response = client.get(f"/sourcing/runs/{run_id}/export.csv", headers=auth_headers())

        assert response.status_code == 200
        lines = response.text.strip().splitlines()
        assert len(lines) == 1  # header row only

    def test_requires_auth(self, client):
        run_id = client.repo.record_sourcing_run(brief_text="find winch manufacturers", target_count=5)
        response = client.get(f"/sourcing/runs/{run_id}/export.csv")
        assert response.status_code == 401


class TestPipelineJobExportEndpoints:
    """GET /pipeline/jobs/{id}/export.csv|.xlsx -- the Find Suppliers
    "Show results" step: every REAL match a run produced (new AND
    duplicate -- a duplicate is a successful, name-corroborated
    validation against an already-existing supplier, not a failure),
    tracker-format, same columns/row-order the category tracker
    exports use (build_tracker_export is category-agnostic)."""

    def test_exports_both_new_and_duplicate_suppliers(self, client):
        new_id = client.repo.create_golden_record({"canonical_name": "Brand New Co", "country": "China"})
        existing_id = client.repo.create_golden_record({"canonical_name": "Already Known Co", "country": "China"})
        client.repo.create_pipeline_job(job_id="job1", query="[discovery] trailer axle", options={})
        client.repo.mark_pipeline_job_completed("job1", stats={
            "new_supplier_ids": [new_id], "duplicate_supplier_ids": [existing_id],
        })

        response = client.get("/pipeline/jobs/job1/export.csv", headers=auth_headers())

        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "Brand New Co" in response.text
        assert "Already Known Co" in response.text

    def test_dedupes_a_supplier_id_appearing_in_both_lists(self, client):
        """Real case: a later round's duplicate can merge into a
        supplier THIS SAME run already created -- the id then appears
        in both new_supplier_ids and duplicate_supplier_ids, but must
        only ever appear once in the export."""
        supplier_id = client.repo.create_golden_record({"canonical_name": "Shared Co", "country": "China"})
        client.repo.create_pipeline_job(job_id="job2", query="[discovery] forklift", options={})
        client.repo.mark_pipeline_job_completed("job2", stats={
            "new_supplier_ids": [supplier_id], "duplicate_supplier_ids": [supplier_id],
        })

        response = client.get("/pipeline/jobs/job2/export.csv", headers=auth_headers())

        assert response.text.count("Shared Co") == 1

    def test_single_company_job_exports_its_resolved_supplier(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Solo Co", "country": "China"})
        client.repo.create_pipeline_job(job_id="job3", query="[single-company] Solo Co", options={})
        client.repo.mark_pipeline_job_completed("job3", stats={"resolved_supplier_id": supplier_id})

        response = client.get("/pipeline/jobs/job3/export.csv", headers=auth_headers())

        assert "Solo Co" in response.text

    def test_sourcing_job_exports_its_qualified_suppliers(self, client):
        """SourcingOutcome uses qualified_supplier_ids, not
        new_supplier_ids/duplicate_supplier_ids -- a sourcing/brief-
        search job's own stats shape, distinct from a discovery job's."""
        id_a = client.repo.create_golden_record({"canonical_name": "Winch Co A", "country": "China"})
        id_b = client.repo.create_golden_record({"canonical_name": "Winch Co B", "country": "China"})
        client.repo.create_pipeline_job(job_id="job7", query="[sourcing] find 2 winch manufacturers", options={})
        client.repo.mark_pipeline_job_completed("job7", stats={"qualified_supplier_ids": [id_a, id_b]})

        response = client.get("/pipeline/jobs/job7/export.csv", headers=auth_headers())

        assert "Winch Co A" in response.text
        assert "Winch Co B" in response.text

    def test_job_with_no_matches_exports_header_only(self, client):
        client.repo.create_pipeline_job(job_id="job4", query="[discovery] nonexistent widget", options={})
        client.repo.mark_pipeline_job_completed("job4", stats={"new_supplier_ids": []})

        response = client.get("/pipeline/jobs/job4/export.csv", headers=auth_headers())

        lines = response.text.strip().splitlines()
        assert len(lines) == 1  # header row only

    def test_unknown_job_returns_404(self, client):
        response = client.get("/pipeline/jobs/nonexistent/export.csv", headers=auth_headers())
        assert response.status_code == 404
        response = client.get("/pipeline/jobs/nonexistent/export.xlsx", headers=auth_headers())
        assert response.status_code == 404

    def test_xlsx_export_returns_spreadsheet_content_type(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Co", "country": "China"})
        client.repo.create_pipeline_job(job_id="job5", query="[discovery] trailer axle", options={})
        client.repo.mark_pipeline_job_completed("job5", stats={"new_supplier_ids": [supplier_id]})

        response = client.get("/pipeline/jobs/job5/export.xlsx", headers=auth_headers())

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_requires_auth(self, client):
        client.repo.create_pipeline_job(job_id="job6", query="[discovery] trailer axle", options={})
        client.repo.mark_pipeline_job_completed("job6", stats={"new_supplier_ids": []})
        assert client.get("/pipeline/jobs/job6/export.csv").status_code == 401
        assert client.get("/pipeline/jobs/job6/export.xlsx").status_code == 401


class TestVerificationHistoryAndChangeLogEndpoints:

    def test_verification_history_returns_recorded_entries(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme"})
        client.repo.record_verification_history(
            supplier_id=supplier_id, verification_type="ai_cross_check", confidence_score=80, verdict="corroborated",
        )
        response = client.get(f"/suppliers/{supplier_id}/verification-history", headers=auth_headers())
        assert response.status_code == 200
        assert len(response.json()) == 1
        assert response.json()[0]["confidence_score"] == 80

    def test_verification_history_unknown_supplier_returns_404(self, client):
        response = client.get("/suppliers/999999/verification-history", headers=auth_headers())
        assert response.status_code == 404

    def test_change_log_returns_recorded_entries(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme", "primary_phone": "+44 20 1234 5678"})
        client.repo.update_supplier_fields_with_history(
            supplier_id, {"primary_phone": "+44 20 9999 0000"}, changed_by="verification_service",
        )
        response = client.get(f"/suppliers/{supplier_id}/change-log", headers=auth_headers())
        assert response.status_code == 200
        assert len(response.json()) == 1
        assert response.json()[0]["field_name"] == "primary_phone"

    def test_change_log_unknown_supplier_returns_404(self, client):
        response = client.get("/suppliers/999999/change-log", headers=auth_headers())
        assert response.status_code == 404

    def test_endpoints_require_auth(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme"})
        assert client.get(f"/suppliers/{supplier_id}/verification-history").status_code == 401
        assert client.get(f"/suppliers/{supplier_id}/change-log").status_code == 401


class TestExportEndpoint:

    def test_export_returns_csv_content_type(self, client):
        client.repo.create_golden_record({"canonical_name": "Acme", "domain": "acme.example.com"})
        response = client.get("/export/csv", headers=auth_headers())
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "Acme" in response.text

    def test_export_requires_auth(self, client):
        response = client.get("/export/csv")
        assert response.status_code == 401

    def test_excel_export_returns_xlsx_content_type(self, client):
        from io import BytesIO

        from openpyxl import load_workbook

        client.repo.create_golden_record({"canonical_name": "Acme", "domain": "acme.example.com"})
        response = client.get("/export/excel", headers=auth_headers())
        assert response.status_code == 200
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        assert any("Acme" in row for row in rows)

    def test_excel_export_requires_auth(self, client):
        response = client.get("/export/excel")
        assert response.status_code == 401


class TestCORSConfiguration:
    """CORSMiddleware is added once, at module-import time, using
    whatever ALLOWED_ORIGINS config.settings had at that moment --
    monkeypatching it per-test after the module has already loaded
    would not actually change the middleware's configured behaviour.
    So this only asserts CORS is wired in at all, not the exact
    allowed-origin list; that list is better verified by hand once
    against the real deployed ALLOWED_ORIGINS env var."""

    def test_cors_middleware_is_registered(self, client):
        import api.app

        middleware_classes = [m.cls.__name__ for m in api.app.app.user_middleware]
        assert "CORSMiddleware" in middleware_classes

    def test_put_is_an_allowed_cors_method(self, client):
        """Regression test for a real bug found live: allow_methods was
        ["GET", "POST"] only, so the browser's own CORS preflight for
        PUT /audit/suppliers/{id}/verdicts/{criterion} failed before
        the request even reached the endpoint -- reproduced as
        "Not saved: UNREACHABLE" in the Audit tab, not an auth or 4xx
        error, since the browser never sent the real request at all.
        Reads the middleware's own configured kwargs directly (see
        this class's own docstring for why ALLOWED_ORIGINS can't be
        monkeypatched after api.app has already been imported -- the
        same is true of allow_methods, so this checks what's actually
        configured rather than exercising a live preflight against
        whatever origin happens to be allowed in this environment)."""
        import api.app

        cors_middleware = next(m for m in api.app.app.user_middleware if m.cls.__name__ == "CORSMiddleware")
        assert "PUT" in cors_middleware.kwargs["allow_methods"]


class TestAuditVerdictEndpoint:
    """PUT /audit/suppliers/{id}/verdicts/{criterion} -- Audit tab
    Stage 2. Every value asserted here is exactly what the request
    body sent, never something the endpoint derives -- see
    api.app.set_audit_verdict's own docstring."""

    def test_needs_auth(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Co", "domain": "acme.com"})
        response = client.put(f"/audit/suppliers/{supplier_id}/verdicts/A", json={"value": "Pass"})
        assert response.status_code == 401

    def test_sets_a_verdict_and_echoes_it_back(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Co", "domain": "acme.com"})
        response = client.put(
            f"/audit/suppliers/{supplier_id}/verdicts/B", json={"value": "Pass"}, headers=auth_headers(),
        )
        assert response.status_code == 200
        assert response.json()["value"] == "Pass"
        assert client.repo.get_audit_verdicts(supplier_id)["B"]["value"] == "Pass"

    def test_sets_notes(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Co", "domain": "acme.com"})
        response = client.put(
            f"/audit/suppliers/{supplier_id}/verdicts/Notes",
            json={"notes": "Called the factory, confirmed by phone."}, headers=auth_headers(),
        )
        assert response.status_code == 200
        assert response.json()["notes"] == "Called the factory, confirmed by phone."

    def test_unknown_supplier_is_404(self, client):
        response = client.put(
            "/audit/suppliers/999999/verdicts/A", json={"value": "Pass"}, headers=auth_headers(),
        )
        assert response.status_code == 404

    def test_unknown_criterion_is_400(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Co", "domain": "acme.com"})
        response = client.put(
            f"/audit/suppliers/{supplier_id}/verdicts/E", json={"value": "Pass"}, headers=auth_headers(),
        )
        assert response.status_code == 400

    def test_value_outside_the_criterions_allowed_set_is_400(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Co", "domain": "acme.com"})
        response = client.put(
            f"/audit/suppliers/{supplier_id}/verdicts/Qualified", json={"value": "Pass"}, headers=auth_headers(),
        )
        assert response.status_code == 400

    def test_get_audit_supplier_bundle_includes_verdicts(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Acme Co", "domain": "acme.com"})
        client.put(f"/audit/suppliers/{supplier_id}/verdicts/A", json={"value": "Fail"}, headers=auth_headers())

        response = client.get(f"/audit/suppliers/{supplier_id}", headers=auth_headers())
        assert response.status_code == 200
        assert response.json()["verdicts"]["A"]["value"] == "Fail"

    def test_get_audit_supplier_bundle_includes_sourcing_dossier(self, client):
        """Real per-brief procurement-fit fields (sourcing/dossier_generator.py),
        needed so the Find Suppliers results screen can show them for a
        detailed-brief search -- previously not exposed by this bundle
        at all."""
        supplier_id = client.repo.create_golden_record({
            "canonical_name": "Acme Winch Co",
            "sourcing_oem_odm_notes": "Acme asserts OEM capability with in-house tooling.",
            "sourcing_verification_status": "verified",
        })

        response = client.get(f"/audit/suppliers/{supplier_id}", headers=auth_headers())

        assert response.status_code == 200
        dossier = response.json()["sourcing_dossier"]
        assert dossier["oem_odm_notes"] == "Acme asserts OEM capability with in-house tooling."
        assert dossier["verification_status"] == "verified"

    def test_sourcing_dossier_fields_are_none_when_never_run(self, client):
        supplier_id = client.repo.create_golden_record({"canonical_name": "Never Sourced Co"})
        response = client.get(f"/audit/suppliers/{supplier_id}", headers=auth_headers())
        assert response.json()["sourcing_dossier"]["oem_odm_notes"] is None
